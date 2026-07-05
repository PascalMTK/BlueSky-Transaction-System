import uuid
import os
import base64
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from decimal import Decimal, InvalidOperation
from django.db.models import Sum, Count, Q
from django.conf import settings
from core.models import User, Country, Transaction, AgentReport
from core.decorators import agent_required, get_auth_user


# ── Logo base64 (cached at startup for email embedding) ────────────────────
def _get_logo_b64() -> str:
    try:
        logo_path = os.path.join(settings.BASE_DIR, 'static', 'images',
                                 'WhatsApp_Image_2026-01-27_at_11.11.59_PM__1_-removebg-preview.png')
        with open(logo_path, 'rb') as f:
            return base64.b64encode(f.read()).decode()
    except Exception:
        return ''

_LOGO_B64 = _get_logo_b64()


def _send_transaction_email(tx, client_email: str, locale: str = 'fr'):
    """Send a styled HTML confirmation email to the client.
    Returns (True, None) on success or (False, error_message) on failure."""
    if not client_email or '@' not in client_email:
        return False, 'Adresse email invalide.'

    from django.core.mail import EmailMultiAlternatives
    from email.mime.image import MIMEImage

    is_en   = locale == 'en'
    is_send = tx.transaction_type == 'send'

    client_name = (tx.sender_name if is_send else tx.receiver_name) or 'Client'
    first_name  = client_name.split()[0]
    amount_str  = f"{tx.total_amount:,.0f} {tx.currency}"
    base_str    = f"{tx.amount:,.0f} {tx.currency}"
    fee_str     = f"{tx.fee_amount:,.0f} {tx.currency}"
    date_str    = tx.sent_at.strftime('%d/%m/%Y  %H:%M') if tx.sent_at else '—'
    year        = tx.sent_at.year if tx.sent_at else datetime.now().year
    origin_name = tx.origin_country.name      if tx.origin_country      else '—'
    dest_name   = tx.destination_country.name if tx.destination_country else '—'
    origin_flag = tx.origin_country.flag_emoji      if tx.origin_country      else ''
    dest_flag   = tx.destination_country.flag_emoji if tx.destination_country else ''
    agent_name  = tx.agent.name if tx.agent else 'BLUESKY'

    if is_en:
        subject      = f"✅ Transfer confirmed — {tx.transaction_number}"
        hi           = f"Hello {first_name},"
        intro        = (f"Your transfer of <b>{amount_str}</b> has been recorded successfully."
                        if is_send else
                        f"Your withdrawal of <b>{amount_str}</b> has been recorded successfully.")
        lbl_ref      = "Reference"; lbl_type  = "Operation"; lbl_origin = "Origin"
        lbl_dest     = "Destination"; lbl_date = "Date"; lbl_agent  = "Agent"
        lbl_base     = "Amount"; lbl_fee   = "Fees"; lbl_total  = "Total paid"
        type_val     = "Transfer" if is_send else "Withdrawal"
        thanks       = "Thank you for choosing <b>BLUESKY Transactions</b>."
        footer_note  = "Automated notification — please do not reply."
        lbl_summary  = "TRANSACTION SUMMARY"; lbl_financial = "FINANCIAL DETAILS"
    else:
        subject      = f"✅ Transaction confirmée — {tx.transaction_number}"
        hi           = f"Bonjour {first_name},"
        intro        = (f"Votre transfert de <b>{amount_str}</b> a été enregistré avec succès."
                        if is_send else
                        f"Votre retrait de <b>{amount_str}</b> a été enregistré avec succès.")
        lbl_ref      = "Référence"; lbl_type  = "Opération"; lbl_origin = "Origine"
        lbl_dest     = "Destination"; lbl_date = "Date"; lbl_agent  = "Agent"
        lbl_base     = "Montant"; lbl_fee   = "Frais"; lbl_total  = "Total payé"
        type_val     = "Transfert" if is_send else "Retrait"
        thanks       = "Merci de faire confiance à <b>BLUESKY Transactions</b>."
        footer_note  = "Notification automatique — merci de ne pas répondre."
        lbl_summary  = "RÉSUMÉ DE LA TRANSACTION"; lbl_financial = "DÉTAILS FINANCIERS"

    html = f"""<!DOCTYPE html>
<html lang="{locale}">
<head><meta charset="UTF-8"><meta name="viewport" content="width=device-width,initial-scale=1"></head>
<body style="margin:0;padding:0;background:#eef2f7;font-family:Inter,Helvetica,Arial,sans-serif;">
<table width="100%" cellpadding="0" cellspacing="0" style="background:#eef2f7;padding:32px 16px;">
<tr><td align="center">
<table width="100%" cellpadding="0" cellspacing="0" style="max-width:560px;">

<tr><td style="background:linear-gradient(145deg,#002d6e,#0055b3,#0096d6);border-radius:16px 16px 0 0;padding:32px 32px 24px;text-align:center;">
  <div style="width:86px;height:86px;border-radius:50%;background:#fff;display:inline-block;line-height:86px;margin-bottom:16px;box-shadow:0 0 0 5px rgba(255,255,255,.20),0 8px 28px rgba(0,0,0,.30);">
    <img src="cid:bluesky_logo" alt="BLUESKY" width="64" height="64" style="width:64px;height:64px;object-fit:contain;vertical-align:middle;display:inline-block;">
  </div>
  <div style="color:#fff;font-size:24px;font-weight:900;letter-spacing:2px;">BLUESKY</div>
  <div style="color:rgba(255,255,255,.55);font-size:10px;letter-spacing:4px;text-transform:uppercase;margin-top:3px;">Transactions</div>
  <div style="margin-top:18px;"><span style="display:inline-block;background:rgba(255,255,255,.18);border:1px solid rgba(255,255,255,.3);border-radius:50px;padding:7px 20px;color:#fff;font-size:13px;font-weight:700;">&#10003; {subject.split(' — ')[0].replace('✅ ','')}</span></div>
</td></tr>

<tr><td style="background:#ffffff;padding:28px 32px 0;">
  <p style="margin:0 0 6px;color:#0f172a;font-size:17px;font-weight:800;">{hi}</p>
  <p style="margin:0 0 24px;color:#475569;font-size:14px;line-height:1.75;">{intro}</p>
</td></tr>

<tr><td style="background:#ffffff;padding:0 32px 20px;">
  <div style="background:#f0f7ff;border-left:4px solid #0284c7;border-radius:0 8px 8px 0;padding:12px 16px;">
    <span style="color:#64748b;font-size:12px;font-weight:600;text-transform:uppercase;letter-spacing:.6px;">{lbl_ref}&nbsp;</span>
    <span style="color:#0284c7;font-size:14px;font-weight:900;font-family:monospace;">{tx.transaction_number}</span>
  </div>
</td></tr>

<tr><td style="background:#ffffff;padding:0 32px 20px;">
  <div style="font-size:10px;font-weight:800;color:#94a3b8;letter-spacing:1.2px;text-transform:uppercase;margin-bottom:10px;">{lbl_summary}</div>
  <table width="100%" cellpadding="0" cellspacing="0" style="border:1px solid #e8edf2;border-radius:10px;overflow:hidden;font-size:13.5px;">
    <tr style="border-bottom:1px solid #e8edf2;"><td style="padding:10px 14px;color:#64748b;font-weight:600;width:45%;">{lbl_type}</td><td style="padding:10px 14px;color:#0f172a;font-weight:700;text-align:right;">{type_val}</td></tr>
    <tr style="border-bottom:1px solid #e8edf2;background:#fafbfc;"><td style="padding:10px 14px;color:#64748b;font-weight:600;">{lbl_origin}</td><td style="padding:10px 14px;color:#0f172a;text-align:right;">{origin_flag} {origin_name}</td></tr>
    <tr style="border-bottom:1px solid #e8edf2;"><td style="padding:10px 14px;color:#64748b;font-weight:600;">{lbl_dest}</td><td style="padding:10px 14px;color:#0f172a;text-align:right;">{dest_flag} {dest_name}</td></tr>
    <tr style="border-bottom:1px solid #e8edf2;background:#fafbfc;"><td style="padding:10px 14px;color:#64748b;font-weight:600;">{lbl_date}</td><td style="padding:10px 14px;color:#0f172a;text-align:right;">{date_str}</td></tr>
    <tr><td style="padding:10px 14px;color:#64748b;font-weight:600;">{lbl_agent}</td><td style="padding:10px 14px;color:#0f172a;text-align:right;">{agent_name}</td></tr>
  </table>
</td></tr>

<tr><td style="background:#ffffff;padding:0 32px 28px;">
  <div style="font-size:10px;font-weight:800;color:#94a3b8;letter-spacing:1.2px;text-transform:uppercase;margin-bottom:10px;">{lbl_financial}</div>
  <table width="100%" cellpadding="0" cellspacing="0" style="font-size:13.5px;">
    <tr><td style="padding:6px 0;color:#64748b;">{lbl_base}</td><td style="padding:6px 0;color:#0f172a;font-weight:600;text-align:right;">{base_str}</td></tr>
    <tr><td style="padding:6px 0;color:#64748b;">{lbl_fee} ({tx.fee_percentage:.1f}%)</td><td style="padding:6px 0;color:#0f172a;font-weight:600;text-align:right;">{fee_str}</td></tr>
  </table>
  <div style="margin-top:12px;background:linear-gradient(135deg,#0055b3,#0096d6);border-radius:10px;padding:14px 18px;">
    <table width="100%" cellpadding="0" cellspacing="0">
      <tr><td style="color:rgba(255,255,255,.8);font-size:13px;font-weight:700;">{lbl_total}</td><td style="color:#fff;font-size:22px;font-weight:900;text-align:right;">{amount_str}</td></tr>
    </table>
  </div>
</td></tr>

<tr><td style="background:#f8fafc;border-top:1px solid #e8edf2;padding:20px 32px;text-align:center;">
  <p style="margin:0;color:#475569;font-size:13.5px;line-height:1.75;">{thanks}</p>
</td></tr>
<tr><td style="background:#f1f5f9;border-radius:0 0 16px 16px;padding:16px 32px;text-align:center;border-top:1px solid #e2e8f0;">
  <p style="margin:0;color:#94a3b8;font-size:11px;">{footer_note}</p>
  <p style="margin:6px 0 0;color:#cbd5e1;font-size:10.5px;">© {year} BLUESKY Transactions</p>
</td></tr>

</table>
</td></tr>
</table>
</body>
</html>"""

    plain = (
        f"{hi}\n\n{intro.replace('<b>','').replace('</b>','')}\n\n"
        f"{lbl_ref}: {tx.transaction_number}\n"
        f"{lbl_type}: {type_val}\n"
        f"{lbl_origin}: {origin_name}\n"
        f"{lbl_dest}: {dest_name}\n"
        f"{lbl_date}: {date_str}\n"
        f"{lbl_total}: {amount_str}\n\n"
        f"{thanks.replace('<b>','').replace('</b>','')}"
    )

    try:
        msg = EmailMultiAlternatives(
            subject=subject,
            body=plain,
            from_email=settings.DEFAULT_FROM_EMAIL,
            to=[client_email],
        )
        msg.attach_alternative(html, 'text/html')
        msg.mixed_subtype = 'related'
        if _LOGO_B64:
            logo_bytes = base64.b64decode(_LOGO_B64)
            logo_mime  = MIMEImage(logo_bytes, _subtype='png')
            logo_mime.add_header('Content-ID', '<bluesky_logo>')
            logo_mime.add_header('Content-Disposition', 'inline', filename='logo.png')
            msg.attach(logo_mime)
        msg.send(fail_silently=False)
        print(f'[BLUESKY EMAIL] sent to {client_email}')
        return True, None
    except Exception as e:
        print(f'[BLUESKY EMAIL] error: {e}')
        return False, str(e)


# ── SMS helper ─────────────────────────────────────────────────────────────

def _build_sms_message(tx, locale='fr'):
    """Return bilingual SMS text for the client after a transaction."""
    amount_str = f"{tx.total_amount:,.0f} {tx.currency}"
    date_str   = tx.sent_at.strftime('%d/%m/%Y %H:%M') if tx.sent_at else ''

    if locale == 'en':
        if tx.transaction_type == 'send':
            name = tx.sender_name or 'Client'
            return (
                f"Hello {name},\n"
                f"Your transfer of {amount_str} (ref: {tx.transaction_number}) "
                f"has been successfully recorded on {date_str}.\n"
                f"Thank you for trusting BLUESKY Transactions!"
            )
        else:
            name = tx.receiver_name or 'Client'
            return (
                f"Hello {name},\n"
                f"Your withdrawal of {amount_str} (ref: {tx.transaction_number}) "
                f"has been successfully recorded on {date_str}.\n"
                f"Thank you for trusting BLUESKY Transactions!"
            )
    else:
        if tx.transaction_type == 'send':
            name = tx.sender_name or 'Client'
            return (
                f"Bonjour {name},\n"
                f"Votre transfert de {amount_str} (réf: {tx.transaction_number}) "
                f"a été enregistré avec succès le {date_str}.\n"
                f"Merci de faire confiance à BLUESKY Transactions !"
            )
        else:
            name = tx.receiver_name or 'Client'
            return (
                f"Bonjour {name},\n"
                f"Votre retrait de {amount_str} (réf: {tx.transaction_number}) "
                f"a été enregistré avec succès le {date_str}.\n"
                f"Merci de faire confiance à BLUESKY Transactions !"
            )

def _send_transaction_sms(tx, locale='fr'):
    """Send SMS to the client. Silent on failure — never blocks the transaction."""
    if not settings.AT_SMS_ENABLED:
        return False
    if not settings.AT_API_KEY:
        return False

    # Pick the right phone number
    phone = None
    if tx.transaction_type == 'send' and tx.sender_phone:
        phone = tx.sender_phone.strip()
    elif tx.transaction_type in ('withdrawal', 'receive') and tx.receiver_phone:
        phone = tx.receiver_phone.strip()

    if not phone:
        return False

    # Normalize phone: ensure it starts with +
    if not phone.startswith('+'):
        phone = '+' + phone.lstrip('0')

    try:
        import africastalking
        africastalking.initialize(settings.AT_USERNAME, settings.AT_API_KEY)
        sms     = africastalking.SMS
        message = _build_sms_message(tx, locale)
        sender  = settings.AT_SENDER_ID or None
        response = sms.send(message, [phone], sender_id=sender)
        print(f'[BLUESKY SMS] sent to {phone}: {response}')
        return True
    except Exception as e:
        print(f'[BLUESKY SMS] error: {e}')
        return False


def _parse_decimal(value, default=0):
    if value is None:
        return float(default or 0)
    if isinstance(value, str):
        value = value.strip()
        if not value:
            return float(default or 0)
    try:
        return float(Decimal(str(value)))
    except (InvalidOperation, TypeError, ValueError):
        return float(default or 0)


def _resolve_fee_amount(post, amount, fallback_fee):
    """The fee is normally derived from a default (country % on create, the
    existing fee on edit), but the agent can edit either the fee or the
    "montant remis" field directly in the UI — whichever was typed in
    'total_amount' takes priority since the client keeps it in sync."""
    total_value = post.get('total_amount', '')
    if total_value is not None and str(total_value).strip() != '':
        total_amt = _parse_decimal(total_value, amount - float(fallback_fee))
        return round(amount - total_amt, 2)
    fee_value = post.get('fee_amount', '')
    if fee_value is None or str(fee_value).strip() == '':
        return round(float(fallback_fee), 2)
    return round(_parse_decimal(fee_value, fallback_fee), 2)


def _build_currencies():
    """Build currency dict from DB countries."""
    currencies = {}
    for c in Country.objects.values('currency_code', 'currency_name').distinct():
        if c['currency_code']:
            currencies[c['currency_code']] = c['currency_name'] or c['currency_code']
    # Add common international currencies not in DB
    extra = {
        'USD': 'Dollar américain', 'EUR': 'Euro', 'GBP': 'Livre sterling',
        'CHF': 'Franc suisse', 'CAD': 'Dollar canadien', 'AED': 'Dirham EAU',
    }
    for k, v in extra.items():
        if k not in currencies:
            currencies[k] = v
    return dict(sorted(currencies.items()))


@agent_required
def dashboard(request):
    user  = get_auth_user(request)
    today = date.today()
    my_tx = Transaction.objects.filter(agent=user)
    stats = {
        'total_transactions': my_tx.count(),
        'transactions_today': my_tx.filter(created_at__date=today).count(),
        'transactions_month': my_tx.filter(created_at__month=today.month, created_at__year=today.year).count(),
        'total_amount':       float(my_tx.filter(status='completed').aggregate(s=Sum('amount'))['s'] or 0),
        'total_fees':         float(my_tx.filter(status='completed').aggregate(s=Sum('fee_amount'))['s'] or 0),
        'amount_month':       float(my_tx.filter(created_at__month=today.month, created_at__year=today.year, status='completed').aggregate(s=Sum('amount'))['s'] or 0),
    }
    recent_tx = my_tx.select_related('origin_country', 'destination_country').order_by('-created_at')[:10]
    return render(request, 'agent/dashboard.html', {'stats': stats, 'recent_tx': recent_tx, 'auth_user': user})


@agent_required
def tx_index(request):
    user = get_auth_user(request)
    qs   = Transaction.objects.filter(agent=user).select_related('origin_country', 'destination_country').order_by('-created_at')
    q           = request.GET.get('q', '')
    status_f    = request.GET.get('status', '')
    type_f      = request.GET.get('transaction_type', '')
    country_f   = request.GET.get('country_id', '')
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    if q:         qs = qs.filter(Q(transaction_number__icontains=q) | Q(sender_name__icontains=q) | Q(receiver_name__icontains=q) | Q(sender_phone__icontains=q))
    if status_f:  qs = qs.filter(status=status_f)
    if type_f:    qs = qs.filter(transaction_type=type_f)
    if country_f: qs = qs.filter(Q(origin_country_id=country_f) | Q(destination_country_id=country_f))
    if date_from:
        try: qs = qs.filter(created_at__date__gte=date_from)
        except Exception: pass
    if date_to:
        try: qs = qs.filter(created_at__date__lte=date_to)
        except Exception: pass
    countries = Country.objects.filter(is_active=True)
    return render(request, 'agent/transactions/index.html', {
        'transactions':      qs,
        'countries':         countries,
        'q':                 q,
        'status_filter':     status_f,
        'type_filter':       type_f,
        'country_filter':    country_f,
        'date_from':         date_from,
        'date_to':           date_to,
        'auth_user':         user,
    })


@agent_required
def tx_create(request):
    user      = get_auth_user(request)
    countries = Country.objects.filter(is_active=True)
    currencies = _build_currencies()
    if request.method == 'POST':
        tx_type   = request.POST.get('transaction_type') or request.GET.get('type', 'send')
        origin_id = request.POST.get('origin_country_id')
        dest_id   = request.POST.get('destination_country_id')
        try:
            origin  = Country.objects.get(pk=origin_id)
            dest    = Country.objects.get(pk=dest_id)
            amount = _parse_decimal(request.POST.get('amount', 0), 0)
            default_pct = float(origin.default_fee_percentage or 0)
            default_fee = round(amount * default_pct / 100, 2)
            fee_amt = _resolve_fee_amount(request.POST, amount, default_fee)
            if fee_amt < 0 or fee_amt > amount:
                messages.error(request, "Les frais ne peuvent pas dépasser le montant donné par le client.")
                raise ValueError('fee_amount_invalid')
            total = round(amount - fee_amt, 2)
            fee_pct = round((fee_amt / amount * 100), 2) if amount else 0
            tx_num  = 'BSK-' + datetime.now().strftime('%Y%m%d') + '-' + str(uuid.uuid4())[:6].upper()
            currency = (request.POST.get('currency') or origin.currency_code or '').upper()
            sent_at_str = request.POST.get('sent_at', '')
            sent_at = None
            if sent_at_str:
                try:
                    sent_at = datetime.fromisoformat(sent_at_str)
                except ValueError:
                    sent_at = None
            if not sent_at:
                sent_at = datetime.now()

            # Validate required fields by type
            sender_name  = request.POST.get('sender_name', '').strip()
            sender_phone = request.POST.get('sender_phone', '').strip()
            recv_name    = request.POST.get('receiver_name', '').strip()
            recv_phone   = request.POST.get('receiver_phone', '').strip()

            if tx_type not in ('send', 'receive', 'exchange', 'withdrawal'):
                tx_type = 'send'
            if tx_type == 'send' and not sender_name:
                messages.error(request, "Le nom de l'expéditeur est obligatoire pour un envoi.")
                raise ValueError('sender_name required')
            if tx_type in ('send', 'exchange', 'withdrawal') and not recv_name:
                messages.error(request, "Le nom du bénéficiaire est obligatoire pour ce type de transaction.")
                raise ValueError('receiver_name required')
            if tx_type == 'exchange' and not sender_name:
                messages.error(request, "Le nom de l'expéditeur est obligatoire pour un échange.")
                raise ValueError('sender_name required')

            tx = Transaction(
                transaction_number   = tx_num,
                transaction_type     = tx_type,
                sender_name          = sender_name,
                sender_phone         = sender_phone,
                receiver_name        = recv_name,
                receiver_phone       = recv_phone,
                client_email         = request.POST.get('client_email', '').strip() or None,
                amount               = amount,
                fee_percentage       = fee_pct,
                fee_amount           = fee_amt,
                total_amount         = total,
                currency             = currency,
                origin_country       = origin,
                destination_country  = dest,
                agent                = user,
                status               = request.POST.get('status', 'completed'),
                notes                = request.POST.get('notes', ''),
                payment_method       = request.POST.get('payment_method', 'cash'),
                sent_at              = sent_at,
            )
            tx.save()
            locale = request.session.get('locale', 'fr')
            _send_transaction_sms(tx, locale)
            messages.success(request, f'Transaction {tx_num} créée avec succès.')
            return redirect('tx_show', tx_id=tx.id)
        except ValueError:
            pass
        except Exception as e:
            messages.error(request, f'Erreur : {e}')
    if request.method == 'POST':
        default_tx_type = request.POST.get('transaction_type', 'send')
        type_preselected = True
    else:
        default_tx_type = request.GET.get('type', '')
        type_preselected = bool(default_tx_type)
        if not default_tx_type:
            default_tx_type = 'send'
    sent_at_value = datetime.now().strftime('%Y-%m-%dT%H:%M')
    return render(request, 'agent/transactions/create.html', {
        'countries': countries, 'currencies': currencies, 'auth_user': user,
        'default_tx_type': default_tx_type,
        'type_preselected': type_preselected,
        'sent_at_value': sent_at_value,
    })


@agent_required
def tx_show(request, tx_id):
    user = get_auth_user(request)
    tx   = get_object_or_404(Transaction, pk=tx_id)
    if not user.is_admin() and tx.agent_id != user.id:
        return redirect('tx_index')
    return render(request, 'agent/transactions/show.html', {'transaction': tx, 'auth_user': user})


@agent_required
def tx_edit(request, tx_id):
    user = get_auth_user(request)
    tx   = get_object_or_404(Transaction, pk=tx_id)
    if not user.is_admin() and tx.agent_id != user.id:
        return redirect('tx_index')
    countries  = Country.objects.filter(is_active=True)
    currencies = _build_currencies()
    if request.method == 'POST':
        tx_type  = request.POST.get('transaction_type', tx.transaction_type)
        amount = _parse_decimal(request.POST.get('amount', tx.amount), tx.amount)
        fee_amt = _resolve_fee_amount(request.POST, amount, tx.fee_amount)
        if fee_amt < 0 or fee_amt > amount:
            messages.error(request, "Les frais ne peuvent pas dépasser le montant donné par le client.")
            return render(request, 'agent/transactions/edit.html', {
                'transaction': tx, 'countries': countries, 'currencies': currencies, 'auth_user': user,
            })
        total = round(amount - fee_amt, 2)
        fee_pct = round((fee_amt / amount * 100), 2) if amount else 0
        currency = (request.POST.get('currency') or tx.currency or '').upper()
        sent_at_str = request.POST.get('sent_at', '')
        if sent_at_str:
            try:
                tx.sent_at = datetime.fromisoformat(sent_at_str)
            except ValueError:
                pass

        tx.transaction_type  = tx_type
        tx.sender_name       = request.POST.get('sender_name', tx.sender_name)
        tx.sender_phone      = request.POST.get('sender_phone', tx.sender_phone)
        tx.receiver_name     = request.POST.get('receiver_name', tx.receiver_name)
        tx.receiver_phone    = request.POST.get('receiver_phone', tx.receiver_phone)
        tx.client_email      = request.POST.get('client_email', '').strip() or None
        tx.amount            = amount
        tx.fee_percentage    = fee_pct
        tx.fee_amount        = fee_amt
        tx.total_amount      = total
        tx.currency          = currency
        tx.status            = request.POST.get('status', tx.status)
        tx.payment_method    = request.POST.get('payment_method', tx.payment_method)
        tx.notes             = request.POST.get('notes', tx.notes)
        origin_id  = request.POST.get('origin_country_id')
        dest_id    = request.POST.get('destination_country_id')
        if origin_id:  tx.origin_country_id = origin_id
        if dest_id:    tx.destination_country_id = dest_id
        tx.save()
        locale = request.session.get('locale', 'fr')
        _send_transaction_sms(tx, locale)
        messages.success(request, 'Transaction mise à jour.')
        return redirect('tx_show', tx_id=tx.id)
    return render(request, 'agent/transactions/edit.html', {
        'transaction': tx, 'countries': countries, 'currencies': currencies, 'auth_user': user,
    })


@agent_required
def tx_send_receipt(request, tx_id):
    """Send (or resend) the confirmation email for a transaction, on demand
    — the agent decides when, rather than it firing automatically on save."""
    user = get_auth_user(request)
    tx   = get_object_or_404(Transaction, pk=tx_id)
    if not user.is_admin() and tx.agent_id != user.id:
        return JsonResponse({'ok': False, 'error': "Vous n'avez pas accès à cette transaction."}, status=403)
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Méthode non autorisée.'}, status=405)

    email = request.POST.get('client_email', '').strip()
    locale = request.session.get('locale', 'fr')
    sent, error = _send_transaction_email(tx, email, locale)
    if sent:
        if tx.client_email != email:
            tx.client_email = email
            tx.save(update_fields=['client_email'])
        return JsonResponse({'ok': True})
    return JsonResponse({'ok': False, 'error': error or 'Échec de l\'envoi.'})


@agent_required
def tx_print(request, tx_id):
    user = get_auth_user(request)
    tx   = get_object_or_404(Transaction, pk=tx_id)
    if not user.is_admin() and tx.agent_id != user.id:
        return redirect('tx_index')
    return render(request, 'agent/transactions/print.html', {'transaction': tx, 'auth_user': user})


@agent_required
def tx_destroy(request, tx_id):
    if request.method == 'POST':
        user = get_auth_user(request)
        tx   = get_object_or_404(Transaction, pk=tx_id)
        if user.is_admin() or tx.agent_id == user.id:
            tx.delete()
            messages.success(request, 'Transaction supprimée.')
    return redirect('tx_index')


@agent_required
def report_store(request):
    if request.method == 'POST':
        user = get_auth_user(request)
        AgentReport.objects.create(
            agent   = user,
            subject = request.POST.get('subject', ''),
            message = request.POST.get('message', ''),
        )
        messages.success(request, 'Rapport envoyé à l\'administrateur.')
    return redirect('agent_dashboard')


@agent_required
def agent_reports_portal(request):
    user = get_auth_user(request)
    if request.method == 'POST':
        subject = request.POST.get('subject', '').strip()
        message = request.POST.get('message', '').strip()
        if subject and message:
            report = AgentReport.objects.create(
                agent   = user,
                subject = subject,
                message = message,
            )
            return redirect(f'/agent/reports/portal/?r={report.id}')
        return redirect('agent_reports_portal')

    reports     = AgentReport.objects.filter(agent=user).order_by('-created_at')
    selected_id = request.GET.get('r')
    selected    = None
    if selected_id:
        try:
            selected = reports.get(id=int(selected_id))
        except (AgentReport.DoesNotExist, ValueError):
            pass

    replied_count = reports.filter(admin_reply__isnull=False).exclude(admin_reply='').count()

    return render(request, 'agent/reports.html', {
        'reports':       reports,
        'selected':      selected,
        'replied_count': replied_count,
        'auth_user':     user,
    })


@agent_required
def report_delete(request, report_id):
    if request.method == 'POST':
        user = get_auth_user(request)
        try:
            report = AgentReport.objects.get(pk=report_id, agent=user)
            report.delete()
        except AgentReport.DoesNotExist:
            pass
    return redirect('agent_reports_portal')


@agent_required
def export_csv(request):
    user = get_auth_user(request)
    qs   = Transaction.objects.filter(agent=user).select_related('origin_country', 'destination_country').order_by('-created_at')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    if date_from:
        try: qs = qs.filter(created_at__date__gte=date_from)
        except Exception: pass
    if date_to:
        try: qs = qs.filter(created_at__date__lte=date_to)
        except Exception: pass

    headers = [
        'N° Transaction', 'Date', 'Type', 'Expéditeur', 'Tél. Expéditeur',
        'Bénéficiaire', 'Tél. Bénéficiaire', 'Montant', 'Frais %',
        'Frais', 'Total', 'Devise', 'Pays Origine', 'Pays Destination', 'Statut',
    ]
    col_widths = [20, 18, 12, 22, 16, 22, 16, 14, 10, 14, 14, 10, 18, 18, 12]
    money_cols = {8, 10, 11}
    text_cols  = {4, 6}

    # Shared styles helper (inline to avoid cross-file dependency)
    thin   = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center')
    left_a = Alignment(horizontal='left',   vertical='center')
    hdr_fill  = PatternFill('solid', fgColor='0284C7')
    even_fill = PatternFill('solid', fgColor='EFF6FF')
    odd_fill  = PatternFill('solid', fgColor='FFFFFF')
    money_fill= PatternFill('solid', fgColor='DBEAFE')

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Mes Transactions — {user.name[:18]}'

    # Header
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = hdr_fill
        cell.font = Font(bold=True, color='FFFFFF', size=11)
        cell.border = border
        cell.alignment = center
        ws.column_dimensions[get_column_letter(col_idx)].width = col_widths[col_idx - 1]
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 28

    TYPE_MAP = {'send': 'Envoi', 'receive': 'Réception', 'exchange': 'Échange', 'withdrawal': 'Retrait'}
    STA_MAP  = {'completed': 'Complété', 'pending': 'En attente', 'cancelled': 'Annulé'}

    # Status cell colors
    sta_colors = {'completed': ('DCFCE7', '15803D'), 'pending': ('FEF9C3', 'CA8A04'), 'cancelled': ('FEE2E2', 'DC2626')}

    for row_idx, t in enumerate(qs, 2):
        row = [
            t.transaction_number,
            t.created_at.strftime('%d/%m/%Y %H:%M'),
            TYPE_MAP.get(t.transaction_type, t.transaction_type),
            t.sender_name,
            t.sender_phone or '',
            t.receiver_name or '',
            t.receiver_phone or '',
            float(t.amount),
            float(t.fee_percentage),
            float(t.fee_amount),
            float(t.total_amount),
            t.currency or '',
            t.origin_country.name,
            t.destination_country.name,
            STA_MAP.get(t.status, t.status),
        ]
        fill = even_fill if row_idx % 2 == 0 else odd_fill
        for col_idx, val in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.border = border
            if col_idx in money_cols:
                cell.fill = money_fill
                cell.number_format = '#,##0.00'
                cell.alignment = center
            elif col_idx == len(headers):   # Statut
                bg, fg = sta_colors.get(t.status, ('FFFFFF', '1E293B'))
                cell.fill = PatternFill('solid', fgColor=bg)
                cell.font = Font(bold=True, color=fg, size=10)
                cell.alignment = center
            elif col_idx in text_cols:
                cell.fill = fill
                cell.alignment = left_a
            else:
                cell.fill = fill
                cell.alignment = center
        ws.row_dimensions[row_idx].height = 20

    ws.auto_filter.ref = ws.dimensions

    filename = f"MesTransactions_{user.agent_code or 'agent'}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response


# ── Network transactions (all countries, read-only) ──────────────────────────

@agent_required
def tx_network(request):
    user     = get_auth_user(request)
    qs       = Transaction.objects.select_related('agent', 'origin_country', 'destination_country').order_by('-created_at')
    q        = request.GET.get('q', '')
    status_f = request.GET.get('status', '')
    country_f= request.GET.get('country_id', '')
    agent_f  = request.GET.get('agent_id', '')
    date_from= request.GET.get('date_from', '')
    date_to  = request.GET.get('date_to', '')
    if q:         qs = qs.filter(Q(transaction_number__icontains=q) | Q(sender_name__icontains=q) | Q(receiver_name__icontains=q))
    if status_f:  qs = qs.filter(status=status_f)
    if country_f: qs = qs.filter(Q(origin_country_id=country_f) | Q(destination_country_id=country_f))
    if agent_f:   qs = qs.filter(agent_id=agent_f)
    if date_from:
        try: qs = qs.filter(created_at__date__gte=date_from)
        except Exception: pass
    if date_to:
        try: qs = qs.filter(created_at__date__lte=date_to)
        except Exception: pass
    countries = Country.objects.filter(is_active=True)
    agents    = User.objects.filter(role='agent', status='active').select_related('country').order_by('name')
    return render(request, 'agent/transactions/network.html', {
        'transactions':  qs[:500],
        'countries':     countries,
        'agents':        agents,
        'q':             q,
        'status_filter': status_f,
        'country_filter':country_f,
        'agent_filter':  agent_f,
        'date_from':     date_from,
        'date_to':       date_to,
        'auth_user':     user,
    })


@agent_required
def fee_for_country(request, country_id):
    try:
        c = Country.objects.get(pk=country_id)
        return JsonResponse({'fee': float(c.default_fee_percentage), 'currency': c.currency_code})
    except Country.DoesNotExist:
        return JsonResponse({'fee': 3.0, 'currency': ''})
