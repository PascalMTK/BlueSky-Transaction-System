import csv, io, json
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime, date
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from django.http import HttpResponse
from django.db.models import Sum, Count, Q, Max
from django.conf import settings
from django.core.mail import send_mail
from core.models import User, Country, Transaction, AgentReport
from core.decorators import admin_required, get_auth_user
from core.views.profile_views import save_profile_photo
from core.mailer import send_async

# Colors inspired by each country's national flag — used to give every
# per-country stat (dashboard breakdown, statistics page) a consistent,
# recognizable color instead of plain black text.
COUNTRY_COLORS = {
    # West Africa
    'CI': '#F77F00',  # Ivory Coast — orange
    'SN': '#00853F',  # Senegal — green
    'ML': '#CE1126',  # Mali — red
    'GN': '#FCD116',  # Guinea — yellow
    'BF': '#EF2B2D',  # Burkina Faso — red
    'GH': '#006B3F',  # Ghana — green
    'NG': '#008751',  # Nigeria — green
    'BJ': '#008751',  # Benin — green
    'TG': '#D21034',  # Togo — red
    'NE': '#E05206',  # Niger — orange
    'MR': '#006233',  # Mauritania — green
    'LR': '#BF0A30',  # Liberia — red
    'SL': '#1EB53A',  # Sierra Leone — green
    'GM': '#3A7728',  # Gambia — green
    'GW': '#CE1126',  # Guinea-Bissau — red
    'CV': '#003893',  # Cape Verde — dark blue
    # Central Africa
    'CD': '#007FFF',  # DRC Congo — sky blue
    'CG': '#009543',  # Congo-Brazzaville — green
    'CM': '#007A5E',  # Cameroon — forest green
    'GA': '#009E60',  # Gabon — green
    'GQ': '#3E9A00',  # Equatorial Guinea — green
    'CF': '#003082',  # Central African Republic — blue
    'TD': '#002664',  # Chad — dark blue
    'ST': '#12AD2B',  # São Tomé — green
    # East Africa
    'KE': '#006600',  # Kenya — dark green
    'TZ': '#1EB53A',  # Tanzania — green
    'UG': '#FCDC04',  # Uganda — yellow
    'RW': '#20603D',  # Rwanda — dark green
    'BI': '#CE1126',  # Burundi — red
    'ET': '#078930',  # Ethiopia — green
    'SO': '#4189DD',  # Somalia — blue
    'DJ': '#6AB2E7',  # Djibouti — light blue
    'ER': '#4189DD',  # Eritrea — blue
    'SS': '#078930',  # South Sudan — green
    # Southern Africa
    'ZA': '#007A4D',  # South Africa — green
    'ZW': '#006400',  # Zimbabwe — dark green
    'ZM': '#198A00',  # Zambia — green
    'MW': '#C61B2D',  # Malawi — red (black removed, invisible on dark bg)
    'MZ': '#009A44',  # Mozambique — green
    'MG': '#FC3D32',  # Madagascar — red
    'AO': '#CC0000',  # Angola — red
    'NA': '#003580',  # Namibia — blue
    'BW': '#75AADB',  # Botswana — blue
    'LS': '#009543',  # Lesotho — green
    'SZ': '#3E5EB9',  # Eswatini — blue
    'MU': '#EA2839',  # Mauritius — red
    'SC': '#003F87',  # Seychelles — blue
    'KM': '#3A75C4',  # Comoros — blue
    # North Africa
    'MA': '#C1272D',  # Morocco — red
    'DZ': '#006233',  # Algeria — green
    'TN': '#E70013',  # Tunisia — red
    'EG': '#C09300',  # Egypt — gold
    'LY': '#239E46',  # Libya — green
    'SD': '#D21034',  # Sudan — red
    # Europe
    'FR': '#0055A4',  # France — blue
    'GB': '#CF142B',  # UK — red
    'BE': '#FAE042',  # Belgium — yellow
    'DE': '#FFCE00',  # Germany — gold
    'PT': '#006600',  # Portugal — green
    'ES': '#AA151B',  # Spain — red
    'IT': '#009246',  # Italy — green
    'CH': '#FF0000',  # Switzerland — red
    'NL': '#AE1C28',  # Netherlands — red
    # Americas
    'US': '#B22234',  # USA — red
    'CA': '#FF0000',  # Canada — red
    'BR': '#009C3B',  # Brazil — green
    # Asia & Middle East
    'CN': '#DE2910',  # China — red
    'AE': '#00732F',  # UAE — green
    'SA': '#006C35',  # Saudi Arabia — green
    'IN': '#FF9933',  # India — orange
    'JP': '#BC002D',  # Japan — red
}
DEFAULT_PALETTE = [
    '#0284c7','#7c3aed','#14b8a6','#f59e0b','#ef4444',
    '#22c55e','#f97316','#8b5cf6','#06b6d4','#ec4899',
]


def _country_color(code, index):
    return COUNTRY_COLORS.get((code or '').upper(), DEFAULT_PALETTE[index % len(DEFAULT_PALETTE)])


def _fees_in_usd(qs):
    """Sum fee_amount across a Transaction queryset, converting each row's
    origin-country currency to USD via that country's usd_exchange_rate
    (units of local currency per 1 USD) — summing raw fee_amount directly
    would mix CDF/ZMW/KES/etc into one meaningless number."""
    rows = qs.values('origin_country_id').annotate(total=Sum('fee_amount'), rate=Max('origin_country__usd_exchange_rate'))
    usd_total = 0.0
    for r in rows:
        rate = float(r['rate'] or 1)
        usd_total += float(r['total'] or 0) / rate if rate else 0
    return usd_total


def _notify_agent_activated(agent):
    """Send activation email to agent when admin activates their account."""
    subject = '[BLUESKY] ✅ Votre compte a été activé — Vous pouvez vous connecter'
    html_body = f"""<!DOCTYPE html>
<html>
<head><meta charset="UTF-8"></head>
<body style="font-family:Inter,sans-serif;background:#f0f4f8;padding:30px 20px;margin:0;">
  <div style="max-width:520px;margin:0 auto;background:white;border-radius:16px;overflow:hidden;box-shadow:0 8px 40px rgba(0,0,0,.12);">
    <div style="background:linear-gradient(135deg,#22c55e,#15803d);padding:28px 32px;text-align:center;">
      <div style="font-size:40px;margin-bottom:8px;">✅</div>
      <div style="color:white;font-size:20px;font-weight:900;letter-spacing:.5px;">BLUESKY TRANSACTIONS</div>
      <div style="color:rgba(255,255,255,.85);font-size:13px;margin-top:4px;">Compte activé avec succès</div>
    </div>
    <div style="padding:32px;">
      <p style="color:#334155;font-size:15px;margin:0 0 16px;">Bonjour <strong>{agent.name}</strong>,</p>
      <p style="color:#64748b;font-size:14px;margin:0 0 20px;line-height:1.7;">
        Excellente nouvelle ! Votre compte <strong>BLUESKY Transactions</strong> vient d'être <strong style="color:#16a34a;">activé</strong> par l'administrateur.<br>
        Vous pouvez maintenant vous connecter et commencer à traiter des transactions.
      </p>
      <div style="background:#f0fdf4;border-left:4px solid #22c55e;border-radius:8px;padding:16px 20px;margin-bottom:24px;">
        <div style="font-size:13px;color:#15803d;font-weight:700;margin-bottom:6px;">🎉 Statut : Actif</div>
        <div style="font-size:13px;color:#166534;line-height:1.6;">
          Votre accès à la plateforme est maintenant disponible. Connectez-vous avec votre email et votre mot de passe.
        </div>
      </div>
      <table style="width:100%;border-collapse:collapse;font-size:13px;margin-bottom:24px;">
        <tr style="background:#f8fafc;"><td style="padding:10px 14px;color:#64748b;font-weight:600;">Nom</td><td style="padding:10px 14px;color:#1e293b;">{agent.name}</td></tr>
        <tr><td style="padding:10px 14px;color:#64748b;font-weight:600;">Email de connexion</td><td style="padding:10px 14px;color:#1e293b;">{agent.email}</td></tr>
        <tr style="background:#f8fafc;"><td style="padding:10px 14px;color:#64748b;font-weight:600;">Code agent</td><td style="padding:10px 14px;color:#0284c7;font-family:monospace;font-weight:700;">{agent.agent_code or '—'}</td></tr>
      </table>
      <div style="text-align:center;margin-bottom:20px;">
        <a href="#" style="display:inline-block;padding:14px 32px;background:linear-gradient(135deg,#0099e6,#0052b2);color:white;text-decoration:none;border-radius:10px;font-size:14px;font-weight:700;">
          Se connecter à BLUESKY
        </a>
      </div>
      <p style="color:#94a3b8;font-size:12px;text-align:center;margin:0;">
        Si vous n'êtes pas à l'origine de cette demande, contactez l'administrateur.
      </p>
    </div>
    <div style="background:#f8fafc;border-top:1px solid #e8edf2;padding:16px;text-align:center;">
      <p style="margin:0;font-size:11px;color:#94a3b8;">© 2026 BLUESKY Transactions — Tous droits réservés</p>
    </div>
  </div>
</body>
</html>"""
    plain = f"Bonjour {agent.name},\n\nVotre compte BLUESKY Transactions a été activé.\nVous pouvez maintenant vous connecter.\n\nEmail : {agent.email}\nCode agent : {agent.agent_code or '—'}\n\n— L'équipe BLUESKY Transactions"
    try:
        send_mail(
            subject=subject,
            message=plain,
            from_email=settings.DEFAULT_FROM_EMAIL,
            recipient_list=[agent.email],
            html_message=html_body,
            fail_silently=True,
        )
    except Exception as e:
        print(f'[BLUESKY] Activation notify error: {e}')


@admin_required
def dashboard(request):
    user  = get_auth_user(request)
    today = date.today()

    stats = {
        'total_transactions':  Transaction.objects.count(),
        'transactions_today':  Transaction.objects.filter(created_at__date=today).count(),
        'transactions_month':  Transaction.objects.filter(created_at__month=today.month, created_at__year=today.year).count(),
        'total_amount':        float(Transaction.objects.filter(status='completed').aggregate(s=Sum('amount'))['s'] or 0),
        'amount_today':        float(Transaction.objects.filter(created_at__date=today, status='completed').aggregate(s=Sum('amount'))['s'] or 0),
        'amount_month':        float(Transaction.objects.filter(created_at__month=today.month, created_at__year=today.year, status='completed').aggregate(s=Sum('amount'))['s'] or 0),
        'total_fees':          float(Transaction.objects.filter(status='completed').aggregate(s=Sum('fee_amount'))['s'] or 0),
        'fees_today':          float(Transaction.objects.filter(created_at__date=today, status='completed').aggregate(s=Sum('fee_amount'))['s'] or 0),
        'fees_month':          float(Transaction.objects.filter(created_at__month=today.month, created_at__year=today.year, status='completed').aggregate(s=Sum('fee_amount'))['s'] or 0),
        'total_agents':        User.objects.filter(role='agent').count(),
        'active_agents':       User.objects.filter(role='agent', status='active').count(),
        'pending_agents':      User.objects.filter(role='agent', status='pending').count(),
        'countries_active':    Country.objects.filter(is_active=True).count(),
        'tx_completed':        Transaction.objects.filter(status='completed').count(),
        'tx_pending':          Transaction.objects.filter(status='pending').count(),
        'tx_cancelled':        Transaction.objects.filter(status='cancelled').count(),
    }

    unread_reports = AgentReport.objects.filter(status='unread').select_related('agent')
    recent_tx      = Transaction.objects.select_related('agent', 'origin_country', 'destination_country').order_by('-created_at')[:10]

    top_agents = (
        User.objects.filter(role='agent', status__in=['active', 'pending', 'inactive'])
        .annotate(tx_count=Count('transaction'), tx_amount=Sum('transaction__amount'))
        .order_by('-tx_amount')[:5]
    )

    months = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun', 'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']
    year = today.year
    monthly_qs = (
        Transaction.objects
        .filter(created_at__year=year, status='completed')
        .values('created_at__month')
        .annotate(total=Sum('amount'), fees=Sum('fee_amount'), count=Count('id'))
    )
    monthly_map = {r['created_at__month']: r for r in monthly_qs}

    monthly_data = []
    for m in range(1, 13):
        r = monthly_map.get(m, {})
        monthly_data.append({
            'month':     months[m - 1],
            'month_num': m,
            'total':     float(r.get('total') or 0),
            'fees':      float(r.get('fees') or 0),
            'count':     r.get('count') or 0,
        })

    # Transactions by origin country (donut chart)
    country_qs = (
        Transaction.objects
        .values('origin_country__name', 'origin_country__flag_emoji', 'origin_country__code')
        .annotate(tx_count=Count('id'), tx_amount=Sum('amount'))
        .order_by('-tx_count')[:8]
    )
    country_tx_data = [
        {
            'name':   r['origin_country__name'] or '?',
            'flag':   r['origin_country__flag_emoji'] or '🌍',
            'code':   (r['origin_country__code'] or '').upper(),
            'count':  r['tx_count'],
            'amount': float(r['tx_amount'] or 0),
            'color':  _country_color(r['origin_country__code'], i),
        }
        for i, r in enumerate(country_qs)
    ]

    pending_agents_list = User.objects.filter(role='agent', status='pending').select_related('country').order_by('created_at')

    # Trailing months only (up to the current one) so the sparkline reads as
    # a recent trend rather than trailing off to zero for future months.
    _recent_months = monthly_data[:today.month][-6:]
    volume_trend = [m['total'] for m in _recent_months]
    fees_trend   = [m['fees'] for m in _recent_months]

    return render(request, 'admin/dashboard.html', {
        'stats':               stats,
        'recent_tx':           recent_tx,
        'top_agents':          top_agents,
        'unread_reports_count': unread_reports.count(),
        'unread_reports':      unread_reports[:3],
        'monthly_data':        monthly_data,
        'volume_trend':        volume_trend,
        'fees_trend':          fees_trend,
        'monthly_data_json':   json.dumps(monthly_data),
        'country_tx_json':     json.dumps(country_tx_data),
        'country_tx_data':     country_tx_data,
        'pending_agents_list': pending_agents_list,
        'auth_user':           user,
    })


@admin_required
def agents(request):
    user = get_auth_user(request)
    qs = User.objects.filter(role='agent').exclude(status='deleted').select_related('country').order_by('-created_at')
    q          = request.GET.get('q', '')
    status_f   = request.GET.get('status', '')
    country_f  = request.GET.get('country_id', '')
    if q:        qs = qs.filter(Q(name__icontains=q) | Q(email__icontains=q) | Q(agent_code__icontains=q))
    if status_f: qs = qs.filter(status=status_f)
    if country_f: qs = qs.filter(country_id=country_f)
    countries = Country.objects.filter(is_active=True)
    deleted_agents = User.objects.filter(role='agent', status='deleted').select_related('country').order_by('-updated_at')
    return render(request, 'admin/agents.html', {
        'agents':         qs,
        'deleted_agents': deleted_agents,
        'countries':      countries,
        'q':              q,
        'status_filter':  status_f,
        'country_filter': country_f,
        'auth_user':      user,
    })


@admin_required
def agent_status(request, agent_id):
    if request.method == 'POST':
        agent  = get_object_or_404(User, pk=agent_id, role='agent')
        status = request.POST.get('status')
        if status in ('active', 'inactive', 'pending'):
            was_pending = agent.status == 'pending'
            agent.status = status
            agent.save()
            if status == 'active' and was_pending:
                send_async(_notify_agent_activated, agent)
            messages.success(request, f'Statut de {agent.name} mis à jour.')
    return redirect('admin_agents')


@admin_required
def agent_promote(request, agent_id):
    if request.method == 'POST':
        agent = get_object_or_404(User, pk=agent_id, role='agent')
        agent.role   = 'admin'
        agent.status = 'active'
        agent.save()
        messages.success(request, f'{agent.name} est maintenant administrateur.')
    next_url = request.POST.get('next', 'admin_agents')
    return redirect(next_url)


@admin_required
def agent_set_password(request, agent_id):
    if request.method == 'POST':
        agent      = get_object_or_404(User, pk=agent_id, role='agent')
        new_pw     = request.POST.get('new_password', '')
        confirm_pw = request.POST.get('confirm_password', '')
        if len(new_pw) < 8:
            messages.error(request, 'Le mot de passe doit contenir au moins 8 caractères.')
        elif new_pw != confirm_pw:
            messages.error(request, 'Les mots de passe ne correspondent pas.')
        else:
            import bcrypt as _bcrypt
            agent.password = _bcrypt.hashpw(new_pw.encode(), _bcrypt.gensalt(10)).decode()
            agent.save()
            messages.success(request, f'Mot de passe de {agent.name} modifié avec succès.')
    next_url = request.POST.get('next', '')
    return redirect(next_url if next_url else 'admin_agents')


@admin_required
def agent_edit(request, agent_id):
    agent = get_object_or_404(User, pk=agent_id, role='agent')
    countries = Country.objects.filter(is_active=True).order_by('name')
    if request.method == 'POST':
        name    = request.POST.get('name', '').strip()
        email   = request.POST.get('email', '').strip().lower()
        phone   = request.POST.get('phone', '').strip()
        country_id = request.POST.get('country_id', '')
        status  = request.POST.get('status', agent.status)
        if not name:
            messages.error(request, 'Le nom est requis.')
        elif not email or '@' not in email:
            messages.error(request, 'Email invalide.')
        elif User.objects.filter(email=email).exclude(pk=agent_id).exists():
            messages.error(request, 'Cet email est déjà utilisé.')
        else:
            agent.name   = name
            agent.email  = email
            agent.phone  = phone or None
            agent.status = status if status in ('active','inactive','pending') else agent.status
            agent.country_id = country_id if country_id else None
            agent.save()
            messages.success(request, f'Agent {agent.name} mis à jour.')
            return redirect('admin_agents')
    return render(request, 'admin/agent_edit.html', {
        'agent': agent, 'countries': countries, 'auth_user': get_auth_user(request),
    })


@admin_required
def agent_photo(request, agent_id):
    agent = get_object_or_404(User, pk=agent_id, role='agent')
    if request.method == 'POST':
        error = save_profile_photo(agent, request.FILES.get('photo'))
        if error:
            messages.error(request, error)
        else:
            messages.success(request, f'Photo de {agent.name} mise à jour.')
    return redirect('admin_agent_edit', agent_id=agent.id)


@admin_required
def agent_destroy(request, agent_id):
    if request.method == 'POST':
        user  = get_auth_user(request)
        agent = get_object_or_404(User, pk=agent_id, role='agent')
        if agent.id == user.id:
            messages.error(request, 'Vous ne pouvez pas supprimer votre propre compte.')
            return redirect('admin_agents')
        tx_count = Transaction.objects.filter(agent=agent).count()
        agent.status = 'deleted'
        agent.save()
        messages.success(request, f'{agent.name} archivé. {tx_count} transaction(s) conservée(s).')
    return redirect('admin_agents')


@admin_required
def agent_permanent_delete(request, agent_id):
    if request.method == 'POST':
        user = get_auth_user(request)
        agent = get_object_or_404(User, pk=agent_id, role='agent')
        if agent.id == user.id:
            messages.error(request, 'Vous ne pouvez pas supprimer votre propre compte.')
            return redirect('admin_agents')
        if agent.status != 'deleted':
            messages.error(request, 'Cet agent n’est pas archivé, il ne peut pas être supprimé définitivement.')
            return redirect('admin_agents')
        agent.delete()
        messages.success(request, f'{agent.name} supprimé définitivement.')
    return redirect('admin_agents')


@admin_required
def transactions(request):
    user = get_auth_user(request)
    qs   = Transaction.objects.select_related('agent', 'origin_country', 'destination_country').order_by('-created_at')
    q           = request.GET.get('q', '')
    status_f    = request.GET.get('status', '')
    country_f   = request.GET.get('country', '')
    date_from   = request.GET.get('date_from', '')
    date_to     = request.GET.get('date_to', '')
    if q:         qs = qs.filter(Q(transaction_number__icontains=q) | Q(sender_name__icontains=q) | Q(receiver_name__icontains=q))
    if status_f:  qs = qs.filter(status=status_f)
    if country_f: qs = qs.filter(Q(origin_country__code=country_f) | Q(destination_country__code=country_f))
    if date_from:
        try: qs = qs.filter(created_at__date__gte=date_from)
        except Exception: pass
    if date_to:
        try: qs = qs.filter(created_at__date__lte=date_to)
        except Exception: pass
    countries = Country.objects.filter(is_active=True)
    return render(request, 'admin/transactions.html', {
        'transactions':   qs[:500],
        'countries':      countries,
        'q':              q,
        'status_filter':  status_f,
        'country_filter': country_f,
        'date_from':      date_from,
        'date_to':        date_to,
        'auth_user':      user,
    })


@admin_required
def reports(request):
    user = get_auth_user(request)
    qs   = AgentReport.objects.select_related('agent').order_by('-created_at')
    return render(request, 'admin/reports.html', {'reports': qs, 'auth_user': user})


@admin_required
def report_read(request, report_id):
    if request.method == 'POST':
        r = get_object_or_404(AgentReport, pk=report_id)
        r.status = 'read'
        r.save()
        messages.success(request, 'Rapport marqué comme lu.')
    return redirect('admin_reports')


@admin_required
def report_reply(request, report_id):
    if request.method == 'POST':
        r     = get_object_or_404(AgentReport, pk=report_id)
        reply = request.POST.get('reply', '').strip()
        if reply:
            r.admin_reply = reply
            r.status      = 'read'
            r.replied_at  = datetime.now()
            r.save()
            messages.success(request, 'Réponse envoyée.')
    return redirect('admin_reports')


# ── Countries ──────────────────────────────────────────────────────────────

@admin_required
def countries_index(request):
    user = get_auth_user(request)
    qs   = Country.objects.all().order_by('name')
    return render(request, 'admin/countries/index.html', {'countries': qs, 'auth_user': user})


@admin_required
def countries_create(request):
    user = get_auth_user(request)
    if request.method == 'POST':
        name      = request.POST.get('name', '').strip()
        code      = request.POST.get('code', '').strip().upper()
        cur_code  = request.POST.get('currency_code', '').strip().upper()
        cur_name  = request.POST.get('currency_name', '').strip()
        flag      = request.POST.get('flag_emoji', '').strip() or _code_to_emoji(code)
        phone_code= request.POST.get('phone_code', '').strip()
        fee_pct   = request.POST.get('default_fee_percentage', 3)
        usd_rate  = request.POST.get('usd_exchange_rate') or 1
        is_active = request.POST.get('is_active') == '1'

        if not name or not code or not cur_code:
            messages.error(request, 'Nom, code et devise sont obligatoires.')
        elif Country.objects.filter(code=code).exists():
            messages.error(request, f'Le code pays {code} existe déjà.')
        else:
            c = Country(
                name=name, code=code, currency_code=cur_code, currency_name=cur_name,
                flag_emoji=flag, phone_code=phone_code,
                default_fee_percentage=fee_pct, usd_exchange_rate=usd_rate, is_active=is_active,
            )
            c.save()
            messages.success(request, f'{name} ajouté avec succès.')
            return redirect('admin_countries')
    return render(request, 'admin/countries/create.html', {'auth_user': user})


@admin_required
def countries_edit(request, country_id):
    user = get_auth_user(request)
    c    = get_object_or_404(Country, pk=country_id)
    if request.method == 'POST':
        c.name                   = request.POST.get('name', c.name).strip()
        c.currency_code          = request.POST.get('currency_code', c.currency_code).strip().upper()
        c.currency_name          = request.POST.get('currency_name', c.currency_name).strip()
        flag = request.POST.get('flag_emoji', '').strip()
        c.flag_emoji             = flag or _code_to_emoji(c.code)
        c.phone_code             = request.POST.get('phone_code', c.phone_code).strip()
        c.default_fee_percentage = request.POST.get('default_fee_percentage', c.default_fee_percentage)
        c.usd_exchange_rate      = request.POST.get('usd_exchange_rate') or c.usd_exchange_rate
        c.is_active              = request.POST.get('is_active') == '1'
        c.save()
        messages.success(request, 'Pays mis à jour.')
        return redirect('admin_countries')
    return render(request, 'admin/countries/edit.html', {'country': c, 'auth_user': user})


@admin_required
def countries_toggle(request, country_id):
    if request.method == 'POST':
        c = get_object_or_404(Country, pk=country_id)
        c.is_active = not c.is_active
        c.save()
        msg = 'Pays activé — il apparaît dans les formulaires.' if c.is_active else 'Pays désactivé.'
        messages.success(request, msg)
    return redirect('admin_countries')


@admin_required
def countries_destroy(request, country_id):
    if request.method == 'POST':
        c = get_object_or_404(Country, pk=country_id)
        if Transaction.objects.filter(Q(origin_country=c) | Q(destination_country=c)).exists():
            messages.error(request, 'Ce pays ne peut pas être supprimé car il a des transactions liées.')
        elif User.objects.filter(country=c).exists():
            messages.error(request, 'Ce pays ne peut pas être supprimé car des agents y sont rattachés.')
        else:
            c.delete()
            messages.success(request, 'Pays supprimé.')
    return redirect('admin_countries')


# ── Statistics ─────────────────────────────────────────────────────────────

@admin_required
def statistics(request):
    user    = get_auth_user(request)
    today   = date.today()
    months  = ['Jan', 'Fév', 'Mar', 'Avr', 'Mai', 'Jun', 'Jul', 'Aoû', 'Sep', 'Oct', 'Nov', 'Déc']

    # Yearly summary (all years found in DB)
    from django.db.models.functions import ExtractYear, ExtractMonth
    yearly_raw = (
        Transaction.objects.filter(status='completed')
        .annotate(year=ExtractYear('created_at'))
        .values('year')
        .annotate(total=Count('id'), total_amount=Sum('amount'), total_fees=Sum('fee_amount'))
        .order_by('year')
    )

    yearly_data = []
    prev_amount = None
    for row in yearly_raw:
        growth = None
        if prev_amount is not None and prev_amount > 0:
            growth = round(((float(row['total_amount'] or 0) - float(prev_amount)) / float(prev_amount)) * 100, 1)
        yearly_data.append({
            'year':         row['year'],
            'total':        row['total'],
            'total_amount': float(row['total_amount'] or 0),
            'total_fees':   float(row['total_fees'] or 0),
            'growth':       growth,
        })
        prev_amount = row['total_amount'] or 0

    # Monthly detail for current year
    current_year_monthly = []
    max_amount = 0
    for m in range(1, 13):
        qs = Transaction.objects.filter(created_at__year=today.year, created_at__month=m, status='completed')
        amt = float(qs.aggregate(s=Sum('amount'))['s'] or 0)
        cnt = qs.count()
        current_year_monthly.append({'month': months[m - 1], 'month_num': m, 'amount': amt, 'count': cnt})
        if amt > max_amount:
            max_amount = amt

    # Converted to USD (via each country's admin-set exchange rate) since
    # fee_amount is stored in each country's own currency — summing those
    # raw amounts across countries would mix CDF/ZMW/KES/etc together.
    daily_gains_today = _fees_in_usd(Transaction.objects.filter(created_at__date=today, status='completed'))
    daily_gains_month = _fees_in_usd(Transaction.objects.filter(created_at__month=today.month, created_at__year=today.year, status='completed'))
    daily_gains_total = _fees_in_usd(Transaction.objects.filter(status='completed'))

    # Status breakdown — surfaces money stuck in pending and the cancellation
    # rate, which the completed-only stats above hide entirely.
    total_tx_all = Transaction.objects.count()
    status_breakdown = []
    for code in ('completed', 'pending', 'cancelled'):
        qs  = Transaction.objects.filter(status=code)
        cnt = qs.count()
        status_breakdown.append({
            'code':   code,
            'count':  cnt,
            'amount': float(qs.aggregate(s=Sum('amount'))['s'] or 0),
            'pct':    round((cnt / total_tx_all * 100), 1) if total_tx_all else 0,
        })

    # Per-country breakdown — colored per country (flag color) instead of
    # plain black text, same palette as the dashboard's country panel.
    country_qs = (
        Transaction.objects.filter(status='completed')
        .values('origin_country__name', 'origin_country__flag_emoji', 'origin_country__code')
        .annotate(tx_count=Count('id'), tx_amount=Sum('amount'))
        .order_by('-tx_count')
    )
    completed_total = sum(r['tx_count'] for r in country_qs) or 0
    country_breakdown = [
        {
            'name':   r['origin_country__name'] or '?',
            'flag':   r['origin_country__flag_emoji'] or '🌍',
            'code':   (r['origin_country__code'] or '').upper(),
            'count':  r['tx_count'],
            'amount': float(r['tx_amount'] or 0),
            'pct':    round((r['tx_count'] / completed_total * 100), 1) if completed_total else 0,
            'color':  _country_color(r['origin_country__code'], i),
        }
        for i, r in enumerate(country_qs)
    ]

    return render(request, 'admin/statistics.html', {
        'yearly_data':          yearly_data,
        'current_year_monthly': current_year_monthly,
        'status_breakdown':     status_breakdown,
        'country_breakdown':    country_breakdown,
        'total_tx_all':         total_tx_all,
        'max_amount':           max_amount,
        'current_year':         today.year,
        'daily_gains_today':    daily_gains_today,
        'daily_gains_month':    daily_gains_month,
        'daily_gains_total':    daily_gains_total,
        'auth_user':            user,
    })


# ── System Reset ──────────────────────────────────────────────────────────

@admin_required
def reset_system(request):
    if request.method == 'POST' and request.POST.get('confirm') == '1':
        Transaction.objects.all().delete()
        messages.success(request, 'Système réinitialisé. Toutes les transactions ont été supprimées.')
    return redirect('admin_dashboard')


@admin_required
def reset_by_country(request, country_id):
    if request.method == 'POST':
        c = get_object_or_404(Country, pk=country_id)
        Transaction.objects.filter(
            Q(origin_country=c) | Q(destination_country=c)
        ).delete()
        messages.success(request, f'Transactions du pays {c.name} supprimées.')
    return redirect('admin_dashboard')


# ── Export Excel ──────────────────────────────────────────────────────────

def _make_xlsx_response(filename: str):
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response

def _style_xlsx(ws, headers: list, col_widths: list):
    """Apply header style + alternating rows to a worksheet."""
    # Styles
    hdr_fill  = PatternFill('solid', fgColor='0284C7')      # sky-600
    hdr_font  = Font(bold=True, color='FFFFFF', size=11)
    even_fill = PatternFill('solid', fgColor='EFF6FF')      # blue-50
    odd_fill  = PatternFill('solid', fgColor='FFFFFF')
    money_fill = PatternFill('solid', fgColor='DBEAFE')     # blue-100 for amount cols
    thin = Side(border_style='thin', color='CBD5E1')
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal='center', vertical='center', wrap_text=False)
    left   = Alignment(horizontal='left',   vertical='center')

    # Header row
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill   = hdr_fill
        cell.font   = hdr_font
        cell.border = border
        cell.alignment = center

    # Freeze header row
    ws.freeze_panes = 'A2'
    ws.row_dimensions[1].height = 28

    # Column widths
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    return even_fill, odd_fill, money_fill, border, center, left

def _apply_row_style(ws, row_idx, num_cols, even_fill, odd_fill,
                     money_fill, money_cols, border, center, left, text_cols):
    fill = even_fill if row_idx % 2 == 0 else odd_fill
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill   = money_fill if col_idx in money_cols else fill
        cell.border = border
        cell.alignment = center if col_idx not in text_cols else left
        ws.row_dimensions[row_idx].height = 20


@admin_required
def export_csv(request):
    user = get_auth_user(request)
    qs   = Transaction.objects.select_related('agent', 'origin_country', 'destination_country').order_by('-created_at')
    date_from = request.GET.get('date_from', '')
    date_to   = request.GET.get('date_to', '')
    if date_from:
        try: qs = qs.filter(created_at__date__gte=date_from)
        except Exception: pass
    if date_to:
        try: qs = qs.filter(created_at__date__lte=date_to)
        except Exception: pass

    headers = [
        'N° Transaction', 'Date', 'Type', 'Expéditeur', 'Tél. Expéditeur',
        'Bénéficiaire', 'Tél. Bénéficiaire', 'Montant', 'Frais %',
        'Frais', 'Total', 'Devise', 'Pays Origine', 'Pays Destination',
        'Paiement', 'Statut', 'Agent',
    ]
    col_widths = [20, 18, 12, 22, 16, 22, 16, 14, 10, 14, 14, 10, 18, 18, 16, 12, 20]
    money_cols = {8, 10, 11}   # Amount, Fee, Total
    text_cols  = {4, 6, 17}    # Sender, Beneficiary, Agent

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Transactions BLUESKY'

    even_fill, odd_fill, money_fill, border, center, left = _style_xlsx(ws, headers, col_widths)

    TYPE_MAP = {'send': 'Envoi', 'receive': 'Réception', 'exchange': 'Échange', 'withdrawal': 'Retrait'}
    PAY_MAP  = {'cash': 'Espèces', 'bank_transfer': 'Virement', 'mobile_money': 'Mobile Money', 'card': 'Carte'}
    STA_MAP  = {'completed': 'Complété', 'pending': 'En attente', 'cancelled': 'Annulé'}

    for row_idx, t in enumerate(qs, 2):
        row = [
            t.transaction_number,
            t.created_at.strftime('%d/%m/%Y %H:%M'),
            TYPE_MAP.get(t.transaction_type, t.transaction_type),
            t.sender_name,
            t.sender_phone or '',
            t.receiver_name or '',
            t.receiver_phone or '',
            float(t.amount),
            float(t.fee_percentage),
            float(t.fee_amount),
            float(t.total_amount),
            t.currency or '',
            t.origin_country.name,
            t.destination_country.name,
            PAY_MAP.get(t.payment_method, t.payment_method),
            STA_MAP.get(t.status, t.status),
            t.agent.name if t.agent else '—',
        ]
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        _apply_row_style(ws, row_idx, len(headers), even_fill, odd_fill,
                         money_fill, money_cols, border, center, left, text_cols)

    # Number format for money columns
    for col_idx in money_cols:
        for row_idx in range(2, ws.max_row + 1):
            ws.cell(row=row_idx, column=col_idx).number_format = '#,##0.00'

    ws.auto_filter.ref = ws.dimensions

    filename = f"BLUESKY_Transactions_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = _make_xlsx_response(filename)
    wb.save(response)
    return response


# ── Helpers ───────────────────────────────────────────────────────────────

def _code_to_emoji(code: str) -> str:
    code = code.strip().upper()
    if len(code) != 2:
        return ''
    return ''.join(chr(0x1F1E6 + ord(c) - ord('A')) for c in code)
