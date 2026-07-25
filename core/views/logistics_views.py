import csv
import io
import json
from datetime import datetime, date
from decimal import Decimal, InvalidOperation

import openpyxl
from django.contrib import messages
from django.db.models import Q, Count, Sum
from django.http import JsonResponse
from django.shortcuts import render, redirect, get_object_or_404

from core.decorators import admin_required, agent_required, get_auth_user
from core.geocoding import geocode_address
from core.logistics_parser import parse_smart_paste
from core.models import Client, ClientAddress, ClientNote, Country, Delivery, LogisticsImportBatch, Transaction, User
from core.routing import nearest_neighbor_order
from core.views.admin_views import _make_xlsx_response, _style_xlsx, _apply_row_style

INLINE_EDITABLE_FIELDS = {'status', 'driver_id'}


def _parse_scheduled_at(value):
    value = (value or '').strip()
    if not value:
        return None
    for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%Y-%m-%d'):
        try:
            return datetime.strptime(value, fmt)
        except ValueError:
            continue
    return None


def _parse_amount(raw):
    raw = (raw or '').strip().replace(',', '.')
    if not raw:
        return None
    try:
        return Decimal(raw)
    except InvalidOperation:
        return None


def _apply_payment_fields(delivery, post):
    """Shared amount/payment-method/address-note handling for the admin
    create/edit forms and the driver's own task-update form."""
    delivery.amount = _parse_amount(post.get('amount'))
    payment_method = post.get('payment_method', '').strip()
    delivery.payment_method = payment_method or None
    other = post.get('payment_method_other', '').strip()
    delivery.payment_method_other = other if payment_method == 'other' else None
    delivery.address_note = post.get('address_note', '').strip() or None


def _status_badge(status):
    labels = {
        'pending':    ('⏳ En attente', 'badge-status status-pending'),
        'in_transit': ('🚚 En transit', 'badge-status status-in_transit'),
        'delayed':    ('⚠️ Retardé', 'badge-status status-delayed'),
        'completed':  ('✅ Terminé', 'badge-status status-completed'),
    }
    return labels.get(status, (status, 'badge-status status-pending'))


# ── Admin — Logistics Dashboard ─────────────────────────────────────────────

def _filtered_deliveries(request, task_type_filter=True):
    """Shared date/driver/status/search filtering for the dashboard, the
    history page, and the Excel export — keeps the three views in sync
    so "what you see is what you export"."""
    base_qs = Delivery.objects.select_related('client', 'driver', 'address')

    date_from = request.GET.get('date_from', '').strip()
    date_to   = request.GET.get('date_to', '').strip()
    driver_f  = request.GET.get('driver', '').strip()
    status_f  = request.GET.get('status', '').strip()
    type_f    = request.GET.get('task_type', '').strip() if task_type_filter else ''
    q         = request.GET.get('q', '').strip()
    sort      = request.GET.get('sort', 'asc')

    qs = base_qs
    if date_from:
        qs = qs.filter(scheduled_at__date__gte=date_from)
    if date_to:
        qs = qs.filter(scheduled_at__date__lte=date_to)
    if driver_f:
        qs = qs.filter(driver_id=driver_f)
    if status_f:
        qs = qs.filter(status=status_f)
    if type_f:
        qs = qs.filter(task_type=type_f)
    if q:
        qs = qs.filter(Q(client__name__icontains=q) | Q(client__phone__icontains=q))

    qs = qs.order_by('scheduled_at' if sort == 'asc' else '-scheduled_at')

    filters = {
        'date_from': date_from, 'date_to': date_to, 'driver_filter': driver_f,
        'status_filter': status_f, 'task_type_filter': type_f, 'q': q, 'sort': sort,
    }
    return qs, filters


@admin_required
def dashboard(request):
    user = get_auth_user(request)
    qs, filters = _filtered_deliveries(request, task_type_filter=False)
    drivers = User.objects.filter(role='agent', status='active').order_by('name')

    status_counts = dict(qs.values_list('status').annotate(n=Count('id')))

    return render(request, 'admin/logistics/dashboard.html', {
        'auth_user':  user,
        'deliveries': qs[:500],
        'drivers':    drivers,
        'status_choices': Delivery.STATUS_CHOICES,
        'total_count': sum(status_counts.values()),
        'status_counts': status_counts,
        **filters,
    })


@admin_required
def history(request):
    user = get_auth_user(request)
    qs, filters = _filtered_deliveries(request)
    drivers = User.objects.filter(role='agent', status='active').order_by('name')

    agg = qs.aggregate(total_amount=Sum('amount'))
    status_counts = dict(qs.values_list('status').annotate(n=Count('id')))

    return render(request, 'admin/logistics/history.html', {
        'auth_user':  user,
        'deliveries': qs[:500],
        'drivers':    drivers,
        'status_choices': Delivery.STATUS_CHOICES,
        'task_type_choices': Delivery.TASK_TYPE_CHOICES,
        'total_count': sum(status_counts.values()),
        'status_counts': status_counts,
        'total_amount': agg['total_amount'] or 0,
        **filters,
    })


@admin_required
def export_xlsx(request):
    qs, _filters = _filtered_deliveries(request)

    headers = ['Date/Heure', 'Type', 'Client', 'Téléphone', 'Adresse', 'Chauffeur', 'Montant', 'Paiement', 'Statut', 'Notes']
    col_widths = [18, 12, 24, 16, 32, 20, 12, 16, 14, 30]
    money_cols = {7}
    text_cols = {3, 4, 5, 6, 10}

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Logistique BLUESKY'

    even_fill, odd_fill, money_fill, border, center, left = _style_xlsx(ws, headers, col_widths)

    TYPE_MAP = {'pickup': 'Collecte', 'dropoff': 'Livraison'}
    STATUS_MAP = dict(Delivery.STATUS_CHOICES)
    PAYMENT_MAP = {'cash': 'Espèces'}

    for row_idx, d in enumerate(qs, 2):
        address_display = d.address_note or (d.address.address_line if d.address else '')
        if d.payment_method == 'other':
            payment_display = d.payment_method_other or 'Autre'
        else:
            payment_display = PAYMENT_MAP.get(d.payment_method, '') if d.payment_method else ''
        row = [
            d.scheduled_at.strftime('%d/%m/%Y %H:%M'),
            TYPE_MAP.get(d.task_type, d.task_type),
            d.client.name,
            d.client.phone,
            address_display,
            d.driver.name if d.driver else '—',
            float(d.amount) if d.amount is not None else '',
            payment_display,
            STATUS_MAP.get(d.status, d.status),
            d.notes or '',
        ]
        for col_idx, val in enumerate(row, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)
        _apply_row_style(ws, row_idx, len(headers), even_fill, odd_fill,
                         money_fill, money_cols, border, center, left, text_cols)

    for row_idx in range(2, ws.max_row + 1):
        ws.cell(row=row_idx, column=7).number_format = '#,##0.00'

    ws.auto_filter.ref = ws.dimensions

    filename = f"BLUESKY_Logistique_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    response = _make_xlsx_response(filename)
    wb.save(response)
    return response


@admin_required
def delivery_create(request):
    user = get_auth_user(request)
    if request.method == 'POST':
        client_id = request.POST.get('client_id', '').strip()
        address_id = request.POST.get('address_id', '').strip()
        task_type = request.POST.get('task_type', 'pickup')
        scheduled_at = _parse_scheduled_at(request.POST.get('scheduled_at'))
        driver_id = request.POST.get('driver_id', '').strip()
        notes = request.POST.get('notes', '').strip()

        if not client_id or not scheduled_at:
            messages.error(request, "Client et date/heure sont obligatoires.")
            return redirect('admin_delivery_create')

        client = get_object_or_404(Client, pk=client_id)
        delivery = Delivery(
            client=client,
            address_id=address_id or None,
            task_type=task_type,
            scheduled_at=scheduled_at,
            driver_id=driver_id or None,
            notes=notes or None,
            created_by=user,
        )
        _apply_payment_fields(delivery, request.POST)
        delivery.save()
        messages.success(request, "Tâche de livraison créée.")
        return redirect('admin_logistics_dashboard')

    clients = Client.objects.prefetch_related('addresses').order_by('name')
    drivers = User.objects.filter(role='agent', status='active').order_by('name')
    preselect_client = request.GET.get('client_id', '')
    return render(request, 'admin/logistics/delivery_create.html', {
        'auth_user': user,
        'clients': clients,
        'drivers': drivers,
        'preselect_client': preselect_client,
        'task_type_choices': Delivery.TASK_TYPE_CHOICES,
        'payment_method_choices': Delivery.PAYMENT_METHOD_CHOICES,
    })


@admin_required
def delivery_edit(request, delivery_id):
    user = get_auth_user(request)
    delivery = get_object_or_404(Delivery.objects.select_related('client'), pk=delivery_id)

    if request.method == 'POST':
        address_id = request.POST.get('address_id', '').strip()
        scheduled_at = _parse_scheduled_at(request.POST.get('scheduled_at'))
        if not scheduled_at:
            messages.error(request, "Date/heure invalide.")
            return redirect('admin_delivery_edit', delivery_id=delivery.id)

        delivery.address_id = address_id or None
        delivery.task_type = request.POST.get('task_type', delivery.task_type)
        delivery.scheduled_at = scheduled_at
        delivery.driver_id = request.POST.get('driver_id') or None
        delivery.status = request.POST.get('status', delivery.status)
        delivery.notes = request.POST.get('notes', '').strip() or None
        _apply_payment_fields(delivery, request.POST)
        delivery.save()
        messages.success(request, "Tâche mise à jour.")
        return redirect('admin_logistics_dashboard')

    drivers = User.objects.filter(role='agent', status='active').order_by('name')
    addresses = delivery.client.addresses.all()
    return render(request, 'admin/logistics/delivery_edit.html', {
        'auth_user': user,
        'delivery': delivery,
        'drivers': drivers,
        'addresses': addresses,
        'task_type_choices': Delivery.TASK_TYPE_CHOICES,
        'status_choices': Delivery.STATUS_CHOICES,
        'payment_method_choices': Delivery.PAYMENT_METHOD_CHOICES,
    })


@admin_required
def delivery_inline_update(request, delivery_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Méthode non autorisée.'}, status=405)

    field = request.POST.get('field', '')
    value = request.POST.get('value', '').strip()
    if field not in INLINE_EDITABLE_FIELDS:
        return JsonResponse({'ok': False, 'error': 'Champ non modifiable.'}, status=400)

    delivery = get_object_or_404(Delivery, pk=delivery_id)

    if field == 'status':
        valid_statuses = dict(Delivery.STATUS_CHOICES)
        if value not in valid_statuses:
            return JsonResponse({'ok': False, 'error': 'Statut invalide.'})
        delivery.status = value
        delivery.save(update_fields=['status', 'updated_at'])
        label, badge_class = _status_badge(value)
        return JsonResponse({'ok': True, 'field': 'status', 'value': value, 'label': label, 'badge_class': badge_class})

    if field == 'driver_id':
        if value:
            driver = User.objects.filter(pk=value, role='agent', status='active').first()
            if not driver:
                return JsonResponse({'ok': False, 'error': "Agent introuvable ou inactif."})
            delivery.driver = driver
            label = driver.name
        else:
            delivery.driver = None
            label = "—"
        delivery.save(update_fields=['driver', 'updated_at'])
        return JsonResponse({'ok': True, 'field': 'driver_id', 'value': value, 'label': label})

    return JsonResponse({'ok': False, 'error': 'Champ non modifiable.'}, status=400)


@admin_required
def delivery_destroy(request, delivery_id):
    if request.method == 'POST':
        delivery = get_object_or_404(Delivery, pk=delivery_id)
        delivery.delete()
        messages.success(request, "Tâche supprimée.")
    return redirect('admin_logistics_dashboard')


# ── Admin — Clients ──────────────────────────────────────────────────────────

@admin_required
def clients_index(request):
    user = get_auth_user(request)
    q = request.GET.get('q', '').strip()
    qs = Client.objects.all().order_by('name')
    if q:
        qs = qs.filter(Q(name__icontains=q) | Q(phone__icontains=q) | Q(email__icontains=q))

    total_count = qs.count()
    with_email_count = qs.exclude(email__isnull=True).exclude(email='').count()
    with_address_count = qs.filter(addresses__isnull=False).distinct().count()

    return render(request, 'admin/logistics/clients.html', {
        'auth_user': user,
        'clients': qs[:500],
        'q': q,
        'total_count': total_count,
        'with_email_count': with_email_count,
        'with_address_count': with_address_count,
    })


@admin_required
def client_create(request):
    user = get_auth_user(request)
    if request.method == 'POST':
        name = request.POST.get('name', '').strip()
        phone = request.POST.get('phone', '').strip()
        email = request.POST.get('email', '').strip()
        country_id = request.POST.get('country_id', '').strip()

        if not name or not phone:
            messages.error(request, "Nom et téléphone sont obligatoires.")
            return redirect('admin_logistics_client_create')

        client = Client.objects.create(
            name=name, phone=phone, email=email or None,
            country_id=country_id or None, created_by=user,
        )
        messages.success(request, "Client créé.")
        return redirect('admin_logistics_client_show', client_id=client.id)

    countries = Country.objects.filter(is_active=True)
    return render(request, 'admin/logistics/client_create.html', {'auth_user': user, 'countries': countries})


@admin_required
def client_show(request, client_id):
    user = get_auth_user(request)
    client = get_object_or_404(Client, pk=client_id)
    addresses = client.addresses.all()
    notes = client.notes.select_related('author').all()
    related_transactions = Transaction.objects.filter(
        Q(sender_phone=client.phone) | Q(receiver_phone=client.phone)
    ).order_by('-created_at')[:20]
    deliveries = client.delivery_set.select_related('driver', 'address').order_by('-scheduled_at')[:50]
    return render(request, 'admin/logistics/client_show.html', {
        'auth_user': user,
        'client': client,
        'addresses': addresses,
        'notes': notes,
        'related_transactions': related_transactions,
        'deliveries': deliveries,
    })


@admin_required
def client_edit(request, client_id):
    user = get_auth_user(request)
    client = get_object_or_404(Client, pk=client_id)
    if request.method == 'POST':
        client.name = request.POST.get('name', '').strip()
        client.phone = request.POST.get('phone', '').strip()
        client.email = request.POST.get('email', '').strip() or None
        client.country_id = request.POST.get('country_id') or None
        client.save()
        messages.success(request, "Client mis à jour.")
        return redirect('admin_logistics_client_show', client_id=client.id)

    countries = Country.objects.filter(is_active=True)
    return render(request, 'admin/logistics/client_edit.html', {'auth_user': user, 'client': client, 'countries': countries})


@admin_required
def address_create(request, client_id):
    client = get_object_or_404(Client, pk=client_id)
    if request.method == 'POST':
        label = request.POST.get('label', '').strip() or 'Autre'
        address_line = request.POST.get('address_line', '').strip()
        city = request.POST.get('city', '').strip()
        country_id = request.POST.get('country_id', '').strip()
        lat = request.POST.get('latitude', '').strip()
        lng = request.POST.get('longitude', '').strip()

        if not address_line:
            messages.error(request, "L'adresse est obligatoire.")
            return redirect('admin_logistics_client_show', client_id=client.id)

        addr = ClientAddress(
            client=client, label=label, address_line=address_line,
            city=city or None, country_id=country_id or None,
        )
        if lat and lng:
            addr.latitude = lat
            addr.longitude = lng
            addr.lat_lng_source = 'manual'
        else:
            addr.lat_lng_source = 'unresolved'
        addr.save()

        if addr.lat_lng_source == 'unresolved':
            geo_lat, geo_lon, error = geocode_address(addr.address_line, addr.city or '', addr.country.name if addr.country else '')
            if geo_lat is not None:
                addr.latitude = geo_lat
                addr.longitude = geo_lon
                addr.lat_lng_source = 'geocoded'
                addr.save(update_fields=['latitude', 'longitude', 'lat_lng_source'])
            else:
                messages.warning(request, f"Adresse enregistrée, mais le géocodage automatique a échoué : {error}")

        messages.success(request, "Adresse ajoutée.")
        return redirect('admin_logistics_client_show', client_id=client.id)

    return redirect('admin_logistics_client_show', client_id=client.id)


@admin_required
def address_edit(request, client_id, address_id):
    client = get_object_or_404(Client, pk=client_id)
    addr = get_object_or_404(ClientAddress, pk=address_id, client=client)
    if request.method == 'POST':
        old_address_line = addr.address_line
        addr.label = request.POST.get('label', '').strip() or 'Autre'
        addr.address_line = request.POST.get('address_line', '').strip()
        addr.city = request.POST.get('city', '').strip() or None
        addr.country_id = request.POST.get('country_id') or None
        lat = request.POST.get('latitude', '').strip()
        lng = request.POST.get('longitude', '').strip()
        addr.is_default = request.POST.get('is_default') == 'on'

        if lat and lng:
            addr.latitude = lat
            addr.longitude = lng
            addr.lat_lng_source = 'manual'
        elif addr.address_line != old_address_line:
            addr.latitude = None
            addr.longitude = None
            addr.lat_lng_source = 'unresolved'
        addr.save()

        if addr.lat_lng_source == 'unresolved':
            geo_lat, geo_lon, error = geocode_address(addr.address_line, addr.city or '', addr.country.name if addr.country else '')
            if geo_lat is not None:
                addr.latitude = geo_lat
                addr.longitude = geo_lon
                addr.lat_lng_source = 'geocoded'
                addr.save(update_fields=['latitude', 'longitude', 'lat_lng_source'])
            else:
                messages.warning(request, f"Adresse mise à jour, mais le géocodage automatique a échoué : {error}")

        messages.success(request, "Adresse mise à jour.")
        return redirect('admin_logistics_client_show', client_id=client.id)

    return redirect('admin_logistics_client_show', client_id=client.id)


@admin_required
def address_destroy(request, client_id, address_id):
    if request.method == 'POST':
        addr = get_object_or_404(ClientAddress, pk=address_id, client_id=client_id)
        addr.delete()
        messages.success(request, "Adresse supprimée.")
    return redirect('admin_logistics_client_show', client_id=client_id)


@admin_required
def address_geocode(request, address_id):
    if request.method == 'POST':
        addr = get_object_or_404(ClientAddress, pk=address_id)
        lat, lon, error = geocode_address(addr.address_line, addr.city or '', addr.country.name if addr.country else '')
        if lat is not None:
            addr.latitude = lat
            addr.longitude = lon
            addr.lat_lng_source = 'geocoded'
            addr.save(update_fields=['latitude', 'longitude', 'lat_lng_source'])
            messages.success(request, "Adresse géocodée avec succès.")
        else:
            messages.error(request, f"Échec du géocodage : {error}")
        return redirect('admin_logistics_client_show', client_id=addr.client_id)
    return redirect('admin_logistics_map')


@admin_required
def note_add(request, client_id):
    if request.method == 'POST':
        user = get_auth_user(request)
        client = get_object_or_404(Client, pk=client_id)
        body = request.POST.get('body', '').strip()
        if body:
            ClientNote.objects.create(client=client, author=user, body=body)
            messages.success(request, "Note ajoutée.")
    return redirect('admin_logistics_client_show', client_id=client_id)


# ── Admin — CSV/Excel import ────────────────────────────────────────────────

IMPORT_TARGET_FIELDS = [
    ('ignore', 'Ignorer'),
    ('client_name', 'Nom du client'),
    ('client_phone', 'Téléphone du client'),
    ('client_email', 'Email du client'),
    ('address_label', "Libellé d'adresse"),
    ('address_line', 'Adresse'),
    ('address_city', 'Ville'),
    ('notes', 'Notes'),
    ('delivery_task_type', 'Type de tâche (pickup/dropoff)'),
    ('delivery_scheduled_at', 'Date/heure prévue'),
    ('delivery_driver_name', 'Nom du chauffeur'),
]

MAX_IMPORT_ROWS = 500


@admin_required
def import_upload(request):
    user = get_auth_user(request)
    if request.method == 'POST':
        f = request.FILES.get('file')
        if not f:
            messages.error(request, "Veuillez sélectionner un fichier.")
            return redirect('admin_logistics_import')

        filename = f.name
        ext = filename.rsplit('.', 1)[-1].lower() if '.' in filename else ''

        rows = []
        headers = []
        try:
            if ext == 'csv':
                text = f.read().decode('utf-8-sig')
                reader = csv.DictReader(io.StringIO(text))
                headers = reader.fieldnames or []
                for row in reader:
                    rows.append(row)
                    if len(rows) >= MAX_IMPORT_ROWS:
                        break
                file_type = 'csv'
            elif ext in ('xlsx', 'xlsm'):
                wb = openpyxl.load_workbook(f, data_only=True)
                ws = wb.active
                rows_iter = ws.iter_rows(values_only=True)
                headers = [str(h) if h is not None else '' for h in next(rows_iter)]
                for values in rows_iter:
                    rows.append({headers[i]: (values[i] if i < len(values) else None) for i in range(len(headers))})
                    if len(rows) >= MAX_IMPORT_ROWS:
                        break
                file_type = 'xlsx'
            else:
                messages.error(request, "Format non supporté. Utilisez un fichier .csv ou .xlsx.")
                return redirect('admin_logistics_import')
        except Exception as e:
            messages.error(request, f"Impossible de lire le fichier : {e}")
            return redirect('admin_logistics_import')

        if not rows:
            messages.error(request, "Le fichier ne contient aucune ligne exploitable.")
            return redirect('admin_logistics_import')

        batch = LogisticsImportBatch.objects.create(
            uploaded_by=user,
            original_filename=filename,
            file_type=file_type,
            raw_rows=rows,
            status='pending_mapping',
        )
        return redirect('admin_logistics_import_map', batch_id=batch.id)

    past_batches = LogisticsImportBatch.objects.all()[:20]
    return render(request, 'admin/logistics/import_upload.html', {'auth_user': user, 'past_batches': past_batches})


@admin_required
def import_map(request, batch_id):
    user = get_auth_user(request)
    batch = get_object_or_404(LogisticsImportBatch, pk=batch_id)
    headers = list(batch.raw_rows[0].keys()) if batch.raw_rows else []

    if request.method == 'POST':
        mapping = {}
        used_targets = set()
        for header in headers:
            target = request.POST.get(f'map_{header}', 'ignore')
            if target != 'ignore':
                if target in used_targets:
                    messages.error(request, f"La cible « {dict(IMPORT_TARGET_FIELDS).get(target, target)} » est utilisée plusieurs fois.")
                    return redirect('admin_logistics_import_map', batch_id=batch.id)
                used_targets.add(target)
            mapping[header] = target
        batch.column_mapping = mapping
        batch.save(update_fields=['column_mapping'])
        return _run_import(request, batch)

    sample_rows = batch.raw_rows[:5]
    return render(request, 'admin/logistics/import_map.html', {
        'auth_user': user,
        'batch': batch,
        'headers': headers,
        'sample_rows': sample_rows,
        'target_fields': IMPORT_TARGET_FIELDS,
    })


def _run_import(request, batch):
    mapping = batch.column_mapping or {}
    reverse_map = {v: k for k, v in mapping.items() if v != 'ignore'}

    created_clients = 0
    merged_clients = 0
    created_addresses = 0
    created_notes = 0
    created_deliveries = 0
    skipped_rows = []

    for i, row in enumerate(batch.raw_rows):
        try:
            name = str(row.get(reverse_map.get('client_name', ''), '') or '').strip()
            phone = str(row.get(reverse_map.get('client_phone', ''), '') or '').strip()
            if not name or not phone:
                skipped_rows.append(f"Ligne {i + 1} : nom ou téléphone manquant.")
                continue

            normalized_phone = phone.replace(' ', '').replace('-', '')
            client = Client.objects.filter(phone=normalized_phone).first()
            if client:
                merged_clients += 1
            else:
                email = str(row.get(reverse_map.get('client_email', ''), '') or '').strip() or None
                client = Client.objects.create(name=name, phone=normalized_phone, email=email, created_by=request.auth_user)
                created_clients += 1

            address_line = str(row.get(reverse_map.get('address_line', ''), '') or '').strip()
            address = None
            if address_line:
                address = ClientAddress.objects.create(
                    client=client,
                    label=str(row.get(reverse_map.get('address_label', ''), '') or 'Autre').strip() or 'Autre',
                    address_line=address_line,
                    city=str(row.get(reverse_map.get('address_city', ''), '') or '').strip() or None,
                    lat_lng_source='unresolved',
                )
                created_addresses += 1

            note_body = str(row.get(reverse_map.get('notes', ''), '') or '').strip()
            if note_body:
                ClientNote.objects.create(client=client, author=request.auth_user, body=note_body)
                created_notes += 1

            if 'delivery_scheduled_at' in reverse_map:
                scheduled_raw = str(row.get(reverse_map.get('delivery_scheduled_at', ''), '') or '').strip()
                scheduled_at = _parse_scheduled_at(scheduled_raw)
                if scheduled_at:
                    driver = None
                    driver_name = str(row.get(reverse_map.get('delivery_driver_name', ''), '') or '').strip()
                    if driver_name:
                        driver = User.objects.filter(name__iexact=driver_name, role='agent', status='active').first()
                        if not driver:
                            skipped_rows.append(f"Ligne {i + 1} : chauffeur « {driver_name} » introuvable, laissé non assigné.")
                    task_type = str(row.get(reverse_map.get('delivery_task_type', ''), '') or 'pickup').strip().lower()
                    if task_type not in ('pickup', 'dropoff'):
                        task_type = 'pickup'
                    Delivery.objects.create(
                        client=client, address=address, task_type=task_type,
                        scheduled_at=scheduled_at, driver=driver, created_by=request.auth_user,
                    )
                    created_deliveries += 1
                else:
                    skipped_rows.append(f"Ligne {i + 1} : date/heure de livraison invalide, aucune tâche créée.")
        except Exception as e:
            skipped_rows.append(f"Ligne {i + 1} : erreur inattendue — {e}")

    batch.created_count = created_clients + created_addresses + created_deliveries
    batch.error_log = '\n'.join(skipped_rows) if skipped_rows else None
    batch.status = 'completed'
    batch.save(update_fields=['created_count', 'error_log', 'status'])

    request.session['_last_import_stats'] = {
        'created_clients': created_clients, 'merged_clients': merged_clients,
        'created_addresses': created_addresses, 'created_notes': created_notes,
        'created_deliveries': created_deliveries, 'skipped_count': len(skipped_rows),
    }
    return redirect('admin_logistics_import_result', batch_id=batch.id)


@admin_required
def import_result(request, batch_id):
    user = get_auth_user(request)
    batch = get_object_or_404(LogisticsImportBatch, pk=batch_id)
    stats = request.session.pop('_last_import_stats', None)
    return render(request, 'admin/logistics/import_result.html', {'auth_user': user, 'batch': batch, 'stats': stats})


@admin_required
def smart_paste(request):
    if request.method == 'POST':
        text = request.POST.get('text', '')
        if not text.strip():
            return JsonResponse({'ok': False, 'error': 'Texte vide.'})
        result = parse_smart_paste(text)
        return JsonResponse(result)

    user = get_auth_user(request)
    return render(request, 'admin/logistics/smart_paste.html', {'auth_user': user})


# ── Admin — Map & routing ───────────────────────────────────────────────────

@admin_required
def route_map(request):
    user = get_auth_user(request)
    addresses = ClientAddress.objects.select_related('client').all()
    geocoded = [a for a in addresses if a.has_coordinates()]
    unresolved = [a for a in addresses if not a.has_coordinates()]
    drivers = User.objects.filter(role='agent', status='active').order_by('name')
    return render(request, 'admin/logistics/map.html', {
        'auth_user': user,
        'geocoded_addresses': geocoded,
        'unresolved_addresses': unresolved,
        'drivers': drivers,
    })


def _driver_route_context(driver, day):
    deliveries = Delivery.objects.select_related('client', 'address').filter(
        driver=driver, scheduled_at__date=day,
    ).order_by('scheduled_at')

    stops = []
    not_on_route = []
    for d in deliveries:
        if d.address and d.address.has_coordinates():
            stops.append((float(d.address.latitude), float(d.address.longitude), d.id))
        else:
            not_on_route.append(d)

    ordered = nearest_neighbor_order(stops)
    delivery_by_id = {d.id: d for d in deliveries}
    ordered_deliveries = [{'delivery': delivery_by_id[o['delivery_id']], 'sequence': o['sequence'],
                           'lat': o['lat'], 'lon': o['lon']} for o in ordered]

    return ordered_deliveries, not_on_route


@admin_required
def driver_route(request, driver_id):
    user = get_auth_user(request)
    driver = get_object_or_404(User, pk=driver_id, role='agent')
    day_param = request.GET.get('date', '').strip()
    try:
        day = datetime.strptime(day_param, '%Y-%m-%d').date() if day_param else date.today()
    except ValueError:
        day = date.today()

    ordered_deliveries, not_on_route = _driver_route_context(driver, day)
    return render(request, 'admin/logistics/driver_route.html', {
        'auth_user': user,
        'driver': driver,
        'day': day,
        'ordered_deliveries': ordered_deliveries,
        'not_on_route': not_on_route,
    })


# ── Agent (driver) side ──────────────────────────────────────────────────────

@agent_required
def driver_tasks(request):
    user = get_auth_user(request)
    base_qs = Delivery.objects.select_related('client', 'address').filter(driver=user)
    status_f = request.GET.get('status', '').strip()
    qs = base_qs.filter(status=status_f) if status_f else base_qs
    qs = qs.order_by('scheduled_at')

    status_counts = dict(base_qs.values_list('status').annotate(n=Count('id')))

    return render(request, 'agent/logistics/tasks.html', {
        'auth_user': user,
        'deliveries': qs[:200],
        'status_filter': status_f,
        'status_choices': Delivery.STATUS_CHOICES,
        'total_count': sum(status_counts.values()),
        'status_counts': status_counts,
    })


@agent_required
def driver_transaction_search(request):
    """Live search across every transaction in the system (any agent,
    any country, any status) so a driver can pick the one their new
    delivery/pickup task is for. Deliberately unrestricted in scope, per
    the requirement — narrowed only by the search term and a result cap."""
    q = request.GET.get('q', '').strip()
    if len(q) < 2:
        return JsonResponse({'results': []})

    qs = Transaction.objects.select_related('origin_country', 'destination_country').filter(
        Q(transaction_number__icontains=q) | Q(sender_name__icontains=q) | Q(receiver_name__icontains=q) |
        Q(sender_phone__icontains=q) | Q(receiver_phone__icontains=q)
    ).order_by('-created_at')[:25]

    TYPE_MAP = {'send': 'Envoi', 'receive': 'Réception', 'exchange': 'Échange', 'withdrawal': 'Retrait'}
    results = [{
        'id': t.id,
        'transaction_number': t.transaction_number,
        'type_display': TYPE_MAP.get(t.transaction_type, t.transaction_type),
        'sender_name': t.sender_name,
        'sender_phone': t.sender_phone,
        'receiver_name': t.receiver_name or '',
        'receiver_phone': t.receiver_phone or '',
        'amount': str(t.total_amount),
        'currency': t.currency or (t.origin_country.currency_code if t.origin_country else ''),
        'created_at': t.created_at.strftime('%d/%m/%Y %H:%M'),
        'route': f"{t.origin_country.flag_emoji if t.origin_country else ''} → {t.destination_country.flag_emoji if t.destination_country else ''}",
    } for t in qs]
    return JsonResponse({'results': results})


@agent_required
def driver_task_create(request):
    user = get_auth_user(request)

    if request.method == 'POST':
        tx_id = request.POST.get('transaction_id', '').strip()
        party = request.POST.get('party', '').strip()
        task_type = request.POST.get('task_type', 'pickup')
        scheduled_at = _parse_scheduled_at(request.POST.get('scheduled_at'))

        if not tx_id or party not in ('sender', 'receiver') or not scheduled_at:
            messages.error(request, "Transaction, partie concernée et date/heure sont obligatoires.")
            return redirect('logistics_driver_task_create')

        tx = get_object_or_404(Transaction, pk=tx_id)
        if party == 'sender':
            name, phone = tx.sender_name, tx.sender_phone
        else:
            name, phone = tx.receiver_name, tx.receiver_phone

        if not name or not phone:
            messages.error(request, "Cette transaction n'a pas de coordonnées pour cette partie.")
            return redirect('logistics_driver_task_create')

        normalized_phone = phone.replace(' ', '').replace('-', '')
        client = Client.objects.filter(phone=normalized_phone).first()
        if not client:
            client = Client.objects.create(name=name, phone=normalized_phone, created_by=user)

        delivery = Delivery(
            client=client,
            transaction=tx,
            task_type=task_type,
            scheduled_at=scheduled_at,
            driver=user,
            notes=request.POST.get('notes', '').strip() or None,
            created_by=user,
        )
        _apply_payment_fields(delivery, request.POST)
        delivery.save()
        messages.success(request, "Tâche créée à partir de la transaction.")
        return redirect('logistics_driver_task_show', delivery_id=delivery.id)

    return render(request, 'agent/logistics/task_create.html', {
        'auth_user': user,
        'task_type_choices': Delivery.TASK_TYPE_CHOICES,
        'payment_method_choices': Delivery.PAYMENT_METHOD_CHOICES,
    })


@agent_required
def driver_task_show(request, delivery_id):
    user = get_auth_user(request)
    delivery = get_object_or_404(
        Delivery.objects.select_related('client', 'address'), pk=delivery_id, driver=user,
    )
    if request.method == 'POST':
        _apply_payment_fields(delivery, request.POST)
        delivery.save(update_fields=['address_note', 'amount', 'payment_method', 'payment_method_other', 'updated_at'])
        messages.success(request, "Informations enregistrées.")
        return redirect('logistics_driver_task_show', delivery_id=delivery.id)

    return render(request, 'agent/logistics/task_show.html', {
        'auth_user': user, 'delivery': delivery, 'status_choices': Delivery.STATUS_CHOICES,
        'payment_method_choices': Delivery.PAYMENT_METHOD_CHOICES,
    })


@agent_required
def driver_task_status_update(request, delivery_id):
    if request.method != 'POST':
        return JsonResponse({'ok': False, 'error': 'Méthode non autorisée.'}, status=405)
    user = get_auth_user(request)
    delivery = get_object_or_404(Delivery, pk=delivery_id, driver=user)
    value = request.POST.get('value', '').strip()
    valid_statuses = dict(Delivery.STATUS_CHOICES)
    if value not in valid_statuses:
        return JsonResponse({'ok': False, 'error': 'Statut invalide.'})
    delivery.status = value
    delivery.save(update_fields=['status', 'updated_at'])
    label, badge_class = _status_badge(value)
    return JsonResponse({'ok': True, 'field': 'status', 'value': value, 'label': label, 'badge_class': badge_class})


@agent_required
def driver_route_self(request):
    user = get_auth_user(request)
    day_param = request.GET.get('date', '').strip()
    try:
        day = datetime.strptime(day_param, '%Y-%m-%d').date() if day_param else date.today()
    except ValueError:
        day = date.today()

    ordered_deliveries, not_on_route = _driver_route_context(user, day)
    return render(request, 'agent/logistics/route.html', {
        'auth_user': user,
        'day': day,
        'ordered_deliveries': ordered_deliveries,
        'not_on_route': not_on_route,
    })
